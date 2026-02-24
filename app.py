# -*- coding: utf-8 -*-
"""Shila Restaurant QFD Dashboard - Main Streamlit Application"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import io
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import matplotlib.pyplot as plt
from ml_analyzer import ShilaMLAnalyzer
from config import COLS, LABELS, COLORS, DATA_DIR, REPORTS_DIR, NOTEBOOKLM_DIR, ANTHROPIC_API_KEY, DASHBOARD_PASSWORD
from analyzer import ShilaAnalyzer
from ai_insights import InsightsGenerator, get_api_setup_instructions

# Page Config
st.set_page_config(page_title="Quality Function Deployment Dashboard", 
                   page_icon="Logo.png", 
                   layout="wide", 
                   initial_sidebar_state="collapsed" # We will hide it via CSS too
)  

def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["password"] == DASHBOARD_PASSWORD:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input(
            "Please enter the Manager Password", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password not correct, show input + error.
        st.text_input(
            "Please enter the Manager Password", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 Password incorrect")
        return False
    else:
        # Password correct.
        return True

if not check_password():
    st.stop()  # Do not run the rest of the app if password is wrong

# ==========================================
# 1. Add the CSS Function
# ==========================================

def load_custom_css():
    st.markdown("""
        <style>
            /* IMPORT FONT (Inter) */
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');

            /* GLOBAL SETTINGS */
            html, body, [class*="css"] {
                font-family: 'Inter', sans-serif;
                color: #3C4257; /* Professional Dark Grey */
            }
            
            /* BACKGROUND */
            .stApp {
                background-color: #F7F9FC; /* Light Blue-Grey Background */
            }

            /* HIDE DEFAULT STREAMLIT ELEMENTS */
            [data-testid="stSidebar"] {display: none;} /* Force hide sidebar */
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}

            /* HEADERS */
            h1, h2, h3 {
                color: #1A1F36;
                font-weight: 700;
                letter-spacing: -0.5px;
            }

            /* METRIC CARDS */
            .metric-container {
                background-color: #FFFFFF;
                border: 1px solid #E1E4E8;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 2px 5px rgba(0,0,0,0.02); /* Very subtle shadow */
                text-align: center;
                height: 100%;
                transition: transform 0.2s ease;
            }
            
            .metric-container:hover {
                transform: translateY(-2px);
                box-shadow: 0 5px 15px rgba(0,0,0,0.05);
                border-color: #4CAF50;
            }

            .metric-icon {
                font-size: 1.8rem;
                margin-bottom: 12px;
                opacity: 0.9;
            }

            .metric-value {
                font-size: 2.0rem; /* Large, bold numbers */
                font-weight: 700;
                color: #1A1F36;
                margin-bottom: 4px;
            }

            .metric-label {
                font-size: 0.85rem;
                text-transform: uppercase;
                letter-spacing: 0.8px;
                color: #697386;
                font-weight: 600;
            }

            /* TABS STYLING */
            .stTabs [data-baseweb="tab-list"] {
                gap: 24px;
            }
            .stTabs [data-baseweb="tab"] {
                height: 50px;
                white-space: pre-wrap;
                background-color: transparent;
                border-radius: 4px;
                color: #697386;
                font-weight: 600;
            }
            .stTabs [aria-selected="true"] {
                background-color: #FFFFFF;
                color: #4CAF50;
                border-bottom: 2px solid #4CAF50;
                box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            }

            /* BUTTONS (Top Right) */
            div.stButton > button {
                border-radius: 8px;
                font-weight: 600;
                padding: 0.5rem 1rem;
            }
            
            /* PILL SHAPE BUTTONS */
            div.stButton > button {
                border-radius: 20px !important;
                padding: 0px 20px !important;
                font-size: 14px !important;
                font-weight: 600 !important;
                height: 38px !important;
                border: 1px solid #E2E8F0 !important;
                transition: all 0.3s ease;
            }

            /* Primary (Active) */
            div.stButton > button[kind="primary"] {
                background: linear-gradient(90deg, #0F766E 0%, #0E7490 100%) !important;
                color: white !important;
                border: none !important;
                box-shadow: 0 4px 6px -1px rgba(15, 118, 110, 0.3);
            }

            /* Secondary (Inactive) */
            div.stButton > button[kind="secondary"] {
                background-color: white !important;
                color: #64748B !important;
            }
            div.stButton > button[kind="secondary"]:hover {
                border-color: #0F766E !important;
                color: #0F766E !important;
                background-color: #F0FDFA !important;
            }

            /* LOGO ALIGNMENT FIX */
            [data-testid="stImage"] {
                display: flex;
                align-items: center;
                height: 100%;
            }
            
            /* REMOVE TOP SPACING */
            /* This is the key fix for the "huge distance" */
            .block-container {
                padding-top: 1rem !important; /* Was default ~6rem */
                padding-bottom: 1rem !important;
                margin-top: 0 !important;
            }
            
            /* Hide the default Streamlit header decoration */
            header {visibility: hidden;} 

            /* MENU LINK STYLING */
            /* Makes buttons look like simple nav links */
            .nav-link > button {
                background: transparent !important;
                border: none !important;
                color: #64748B !important;
                font-weight: 600 !important;
                font-size: 15px !important;
                padding: 0 10px !important;
            }
            .nav-link > button:hover {
                color: #0F766E !important; /* Teal on hover */
                background: transparent !important;
            }
            
            /* 1. REMOVE HUGE TOP WHITESPACE */
            .block-container {
                padding-top: 1rem !important;
                padding-bottom: 1rem !important;
                margin-top: 0 !important;
            }
            header {visibility: hidden;}

            /* 2. THE DARK HEADER BACKGROUND */
            /* This creates the dark box behind the menu */
            .block-container {
                padding-top: 1rem !important;
                padding-bottom: 1rem !important;
                margin-top: 0 !important;
            }
            header {visibility: hidden;}
            
            .nav-background {
                position: absolute;
                top: -1rem;
                left: -5rem;
                right: -5rem;
                height: 100px;
                background-color: #0F172A; /* Dark Midnight */
                z-index: 0;
            }

            /* 3. MENU LINKS (Clean Text, No Box) */
            .nav-link > button {
                background-color: transparent !important;
                border: none !important;
                box-shadow: none !important;
                
                /* Font Settings to match Workflow Section */
                color: #FFFFFF !important;
                font-family: sans-serif !important; /* Matches Analysis Workflow */
                font-size: 14px !important;
                font-weight: 700 !important;
                
                /* Layout */
                padding: 0 15px !important;
                margin-top: 6px !important; /* Pushes text down to align with Logo */
                height: auto !important;
            }
            .nav-link > button:hover {
                color: #2DD4BF !important; /* Teal on hover */
            }

            /* Hover Effect (Text turns white) */
            div.stButton > button:hover {
                color: #FFFFFF !important;
                background-color: transparent !important;
            }

            /* Remove default Streamlit focus borders */
            .nav-link > button:active, 
            .nav-link > button:focus {
                background-color: transparent !important;
                border: none !important;
                box-shadow: none !important;
                color: #FFFFFF !important;
            }

            /* 4. LANGUAGE BUTTONS (Dark Mode Style) */
            div.stButton > button {
                background-color: transparent !important;
                border: none !important;
                box-shadow: none !important;
                
                /* Default Text Color (Greyish) */
                color: #94A3B8 !important; 
                font-family: sans-serif !important;
                font-size: 14px !important;
                font-weight: 600 !important;
                padding: 0 10px !important;
                margin-top: 6px !important; /* Aligns with Menu */
            }
            
            div.stButton > button[kind="primary"] {
                color: #FFFFFF !important;
                font-weight: 800 !important;
            }
            div.stButton > button[kind="secondary"] {
                color: #64748B !important; /* Dimmed for inactive */
            }

        </style>
    """, unsafe_allow_html=True)

load_custom_css()

# ==========================================
# 2. Hero layout
# ==========================================

import streamlit.components.v1 as components

def render_hero_section():
    html_code = """
    <!DOCTYPE html>
    <html>
    <head>
        <script src="https://cdn.tailwindcss.com"></script>
        <style>
            /* ANIMATIONS */
            @keyframes fadeInUp { from { opacity: 0; transform: translate3d(0, 20px, 0); } to { opacity: 1; transform: translate3d(0, 0, 0); } }
            @keyframes float { 0%, 100% { transform: translateY(0); } 50% { transform: translateY(-10px); } }
            
            /* Blue Pulse for Processing */
            @keyframes pulseBlue { 0%, 100% { box-shadow: 0 0 0 rgba(59, 130, 246, 0); } 50% { box-shadow: 0 0 15px rgba(59, 130, 246, 0.3); } }
            
            .animate-fade-in-up { animation: fadeInUp 0.8s ease-out forwards; }
            .animate-float { animation: float 6s ease-in-out infinite; }
            .pulse-active { animation: pulseBlue 2s infinite; }
            
            body { margin: 0; overflow: hidden; font-family: sans-serif; background: transparent; }
            .step-card { transition: all 0.5s ease-in-out; }
        </style>
    </head>
    <body>
        <section class="relative py-8 mx-2 overflow-visible">
            
            <div class="absolute inset-0 bg-gradient-to-br from-transparent via-blue-50/20 to-teal-50/20 -z-10"></div>
            <div class="absolute top-0 right-0 w-64 h-64 bg-teal-100 rounded-full mix-blend-multiply filter blur-3xl opacity-30 animate-float -z-10"></div>
            <div class="absolute bottom-0 left-0 w-64 h-64 bg-blue-100 rounded-full mix-blend-multiply filter blur-3xl opacity-30 animate-float -z-10" style="animation-delay: 2s"></div>

            <div class="container mx-auto px-4 relative">
                <div class="grid lg:grid-cols-2 gap-12 items-center">
                    
                    <div class="space-y-6 animate-fade-in-up">
                        <div class="space-y-4">
                            <span class="inline-flex items-center justify-center rounded-full border border-teal-200 bg-teal-50 px-4 py-1 text-xs font-bold text-teal-700 uppercase tracking-wide">
                                AI-Powered Data Analysis
                            </span>
                            <h1 class="text-4xl lg:text-5xl font-extrabold leading-tight text-gray-900 tracking-tight">
                                Stop Guessing. <br/>
                                Turn Feedback into <br/>
                                <span class="text-transparent bg-clip-text bg-gradient-to-r from-teal-500 to-blue-600">Winning Strategies</span>
                            </h1>
                            <p class="text-lg text-gray-600 leading-relaxed max-w-lg">
                                Don't just collect data. <b>Synthesize quantitative metrics with qualitative feedback</b> to turn raw data into actionable insights.                            </p>
                        </div>
                        <div class="flex flex-col sm:flex-row gap-4">
                            <button class="inline-flex items-center justify-center gap-2 font-bold bg-blue-600 text-white shadow-lg shadow-blue-200 hover:bg-blue-700 hover:shadow-xl h-12 rounded-xl px-8 transition-all transform hover:-translate-y-0.5">
                                Start Free
                            </button>
                        </div>
                    </div>

                    <div class="relative animate-fade-in-up" style="animation-delay: 0.2s;">
                        <div class="relative bg-white border border-gray-100 rounded-2xl p-6 shadow-2xl">
                            <div class="text-center mb-6">
                                <h3 class="text-lg font-bold text-gray-900">Analysis Workflow</h3>
                                <p class="text-sm text-gray-500">Live Simulation</p>
                            </div>
                            
                            <div class="space-y-3" id="workflow-container">
                                </div>
                        </div>
                    </div>
                </div>
            </div>
        </section>

        <script>
            // --- 1. DATA CONFIGURATION ---
            const steps = [
                { 
                    id: 1, 
                    title: "Upload Your Data", 
                    desc: "CSV/Excel files", 
                    svg: `<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12" />` 
                },
                { 
                    id: 2, 
                    title: "AI Processing", 
                    desc: "Smart Data Cleaning & Preprocessing", 
                    svg: `<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9.663 17h4.673M12 3v1m6.364 1.636l-.707.707M21 12h-1M4 12H3m3.343-5.657l-.707-.707m2.828 9.9a5 5 0 117.072 0l-.548.547A3.374 3.374 0 0014 18.469V19a2 2 0 11-4 0v-.531c0-.895-.356-1.754-.988-2.386l-.548-.547z" />` 
                },
                { 
                    id: 3, 
                    title: "AI-powered Q&A", 
                    desc: "Query yourdata with natural language processing", 
                    svg: `<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M8 10h.01M12 10h.01M16 10h.01M9 16H5a2 2 0 01-2-2V6a2 2 0 012-2h14a2 2 0 012 2v8a2 2 0 01-2 2h-5l-5 5v-5z" />` 
                },
                { 
                    id: 4, 
                    title: "Get Insights", 
                    desc: "Actionable insights with charts & reports", 
                    svg: `<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z" />` 
                }
            ];

            const container = document.getElementById('workflow-container');

            // --- 2. BUILD HTML STRUCTURE ---
            steps.forEach(step => {
                const html = `
                <div id="step-${step.id}" class="step-card flex items-center p-3 rounded-xl border border-gray-100 bg-white opacity-60">
                    
                    <div class="flex-shrink-0 mr-4">
                        <div id="status-bg-${step.id}" class="w-6 h-6 rounded-full bg-gray-200 flex items-center justify-center text-gray-500 font-bold text-xs transition-colors duration-500">
                            <span id="status-num-${step.id}">${step.id}</span>
                            <svg id="status-check-${step.id}" class="w-3 h-3 text-white hidden" fill="currentColor" viewBox="0 0 20 20"><path fill-rule="evenodd" d="M16.707 5.293a1 1 0 010 1.414l-8 8a1 1 0 01-1.414 0l-4-4a1 1 0 011.414-1.414L8 12.586l7.293-7.293a1 1 0 011.414 0z" clip-rule="evenodd" /></svg>
                        </div>
                    </div>

                    <div class="flex-shrink-0 mr-4">
                        <div id="feat-bg-${step.id}" class="w-10 h-10 rounded-lg bg-gray-100 flex items-center justify-center text-gray-400 transition-colors duration-500">
                            <svg class="w-5 h-5" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                                ${step.svg}
                            </svg>
                        </div>
                    </div>

                    <div class="flex-1 min-w-0">
                        <h4 id="title-${step.id}" class="text-sm font-bold text-gray-500 transition-colors duration-500">${step.title}</h4>
                        <p class="text-xs text-gray-400 transition-colors duration-500 truncate">${step.desc}</p>
                    </div>

                    <span id="badge-${step.id}" class="inline-flex items-center px-2 py-0.5 rounded text-[10px] font-bold bg-gray-100 text-gray-400 transition-all duration-500 uppercase tracking-wide">Wait</span>
                </div>
                `;
                container.innerHTML += html;
            });

            // --- 3. STATE FUNCTIONS ---
            function setIdle(id) {
                const card = document.getElementById(`step-${id}`);
                card.className = "step-card flex items-center p-3 rounded-xl border border-gray-100 bg-white opacity-60";
                updateVisuals(id, "idle");
            }

            function setProcessing(id) {
                const card = document.getElementById(`step-${id}`);
                card.className = "step-card flex items-center p-3 rounded-xl border-2 border-blue-300 bg-blue-50 transform scale-105 shadow-lg opacity-100 pulse-active";
                updateVisuals(id, "processing");
            }

            function setComplete(id) {
                const card = document.getElementById(`step-${id}`);
                card.className = "step-card flex items-center p-3 rounded-xl border border-teal-500 bg-teal-50 opacity-100";
                updateVisuals(id, "complete");
            }

            function updateVisuals(id, state) {
                const statusBg = document.getElementById(`status-bg-${id}`);
                const statusNum = document.getElementById(`status-num-${id}`);
                const statusCheck = document.getElementById(`status-check-${id}`);
                
                const featBg = document.getElementById(`feat-bg-${id}`);
                const title = document.getElementById(`title-${id}`);
                const badge = document.getElementById(`badge-${id}`);

                if (state === 'idle') {
                    statusBg.className = "w-6 h-6 rounded-full bg-gray-200 flex items-center justify-center text-gray-500 font-bold text-xs";
                    statusNum.classList.remove('hidden');
                    statusCheck.classList.add('hidden');
                    featBg.className = "w-10 h-10 rounded-lg bg-gray-100 flex items-center justify-center text-gray-400";
                    title.className = "text-sm font-bold text-gray-500";
                    badge.className = "inline-flex items-center px-2 py-0.5 rounded text-[10px] font-bold bg-gray-100 text-gray-400 uppercase tracking-wide";
                    badge.innerText = "WAIT";
                } 
                else if (state === 'processing') {
                    statusBg.className = "w-6 h-6 rounded-full bg-blue-500 flex items-center justify-center text-white font-bold text-xs animate-pulse";
                    statusNum.classList.remove('hidden');
                    statusCheck.classList.add('hidden');
                    featBg.className = "w-10 h-10 rounded-lg bg-blue-200 flex items-center justify-center text-blue-700 animate-pulse";
                    title.className = "text-sm font-bold text-blue-900";
                    badge.className = "inline-flex items-center px-2 py-0.5 rounded text-[10px] font-bold bg-blue-500 text-white uppercase tracking-wide animate-pulse";
                    badge.innerText = "RUNNING";
                }
                else if (state === 'complete') {
                    statusBg.className = "w-6 h-6 rounded-full bg-teal-500 flex items-center justify-center shadow-sm";
                    statusNum.classList.add('hidden');
                    statusCheck.classList.remove('hidden'); 
                    featBg.className = "w-10 h-10 rounded-lg bg-teal-500 flex items-center justify-center text-white shadow-sm";
                    title.className = "text-sm font-bold text-teal-900";
                    badge.className = "inline-flex items-center px-2 py-0.5 rounded text-[10px] font-bold bg-teal-600 text-white uppercase tracking-wide";
                    badge.innerText = "DONE";
                }
            }

            // --- 4. ANIMATION LOOP ---
            const sleep = m => new Promise(r => setTimeout(r, m));

            async function runLoop() {
                while(true) {
                    steps.forEach(s => setIdle(s.id));
                    await sleep(800);

                    for (let i = 1; i <= 4; i++) {
                        setProcessing(i);
                        await sleep(1500); 
                        setComplete(i);
                        await sleep(300); 
                    }
                    await sleep(3000); 
                }
            }

            runLoop();
        </script>
    </body>
    </html>
    """
    
    components.html(html_code, height=650, scrolling=False)

# ==========================================
# 3. Add the Helper Function
# ==========================================
# ==========================================
# SNAPPFOOD FILE LOADER
# ==========================================

def load_snappfood_file(uploaded_file):
    """Load SnappFood format Excel file and convert to standard format (NPS/Pareto/Kano IGNORED)"""
    from config import SNAPPFOOD_COLS, SNAPPFOOD_ISSUES, COLS
    import pandas as pd
    
    df_raw = pd.read_excel(uploaded_file, sheet_name='Reviews', header=None)
    
    # Find where reviews end
    reviews_end = len(df_raw)
    for i in range(len(df_raw)):
        if i < len(df_raw) and str(df_raw.iloc[i, 1]) == 'Products Rate':
            reviews_end = i
            break
    
    reviews_data = []
    for i in range(3, reviews_end):
        row = df_raw.iloc[i]
        
        # Branch (Anchor B/1)
        branch = str(row[1]).strip() if pd.notna(row[1]) else ""
        if not branch or branch in ['Branch', 'Vendor ID', 'None', 'nan']: continue
        if any(kw in branch for kw in ['Vendor', 'Page']): continue

        # Rating (Anchor O/14)
        try:
            rating = float(row[14]) if pd.notna(row[14]) else None
        except:
            rating = None
        
        # Full Timestamp (Column M / Index 12) - "Order Created At"
        # This contains the '26/12/2025 18:37:39' format
        timestamp_data = row[12] if pd.notna(row[12]) else None

        # Comments (Anchor U/20 and Y/24)
        comment = str(row[20]).strip() if pd.notna(row[20]) else ""
        deliv_comment = str(row[24]).strip() if pd.notna(row[24]) else ""
        full_comment = f"{comment} | {deliv_comment}".strip(" |")
        
        record = {
            COLS['CREATED_AT']: timestamp_data,
            COLS['BRANCH']: branch,
            COLS['RATING']: rating,
            COLS['NPS']: None,       # Ignored
            COLS['WEAKNESS']: None,  # Ignored for Pareto
            COLS['STRENGTH']: None,  # Ignored for Kano
            COLS['ORDER_ITEMS']: str(row[28]) if pd.notna(row[28]) else "",
            COLS['DATE']: timestamp_data,
            COLS['COMMENT']: full_comment if full_comment else None,
            'order_code': row[9],
            'customer_name': str(row[3]).strip() if pd.notna(row[3]) else "Unknown"
        }
        reviews_data.append(record)
    
    # 4. Final Processing
    df = pd.DataFrame(reviews_data)
    
    # Clean up dates immediately
    if not df.empty:
        # Convert the standardized column to datetime
        df[COLS['CREATED_AT']] = pd.to_datetime(df[COLS['CREATED_AT']], dayfirst=True, errors='coerce')
    
    return df

def detect_file_format(uploaded_file):
    """Detect if file is SnappFood format or original format"""
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheets = xl.sheet_names
        
        # SnappFood format has these sheets
        if 'Reviews' in sheets and 'Overview' in sheets:
            return 'snappfood'
        
        # Check for original format columns
        df_check = pd.read_excel(uploaded_file, nrows=3)
        cols = [str(c) for c in df_check.columns]
        
        if any('شعبه' in c for c in cols) or any('میزان رضایت' in c for c in cols):
            return 'original'
        
        return 'original'  # Default
    except:
        return 'original'

def get_metric_html(label, value, icon, color="black"):
    # Detect direction for RTL support
    direction = "rtl" if st.session_state.lang == 'fa' else "ltr"
    
    return f"""
    <div class="metric-card">
        <div class="metric-icon">{icon}</div>
        <div class="metric-label">{label}</div>
        <div class="metric-value" style="color: {color}">{value}</div>
    </div>
    """
def L(key): return LABELS[st.session_state.lang].get(key, key)

def load_data(file):
    try:
        if isinstance(file, str):
            return pd.read_csv(file) if file.endswith('.csv') else pd.read_excel(file)
        return pd.read_csv(file) if file.name.endswith('.csv') else pd.read_excel(file)
    except Exception as e:
        st.error(f"Error: {e}")
        return None

# Session State
if 'lang' not in st.session_state: st.session_state.lang = 'en'
if 'df' not in st.session_state: st.session_state.df = None
if 'analyzer' not in st.session_state: st.session_state.analyzer = None

# ==========================================
# 4. TOP NAVIGATION BAR
# ==========================================

# A. INJECT DARK BACKGROUND
# This div uses the CSS class .nav-background to create the black bar
st.markdown('<div class="nav-background"></div>', unsafe_allow_html=True)

# B. LAYOUT COLUMNS
# Logo (3) | Menu (5) | Language (2)
col_nav_1, col_nav_2, col_nav_3 = st.columns([3, 5, 2])

# 1. LOGO (Top Left)
with col_nav_1:
    try:
        # width=220 is a sweet spot for "InsightForge" text logos
        st.image("Logo.png", width=220) 
    except:
        # Fallback text in White if image fails
        st.markdown("<h2 style='color: white; margin:0;'>InsightForge</h2>", unsafe_allow_html=True)

# 2. MENU ITEMS (Middle)
with col_nav_2:
    # We use a container to help with vertical alignment if needed
    with st.container():
        # Adjust column ratios to control spacing between items
        # [0.5, 1, 1, 1, 0.5] centers them with even spacing
        _, m1, m2, m3, _ = st.columns([0.5, 1, 1, 1, 0.5])
        
        with m1:
            st.markdown('<div class="nav-link">', unsafe_allow_html=True)
            if st.button("Services", key="nav_services", use_container_width=True):
                st.info("Services clicked")
            st.markdown('</div>', unsafe_allow_html=True)
            
        with m2:
            st.markdown('<div class="nav-link">', unsafe_allow_html=True)
            if st.button("Pricing", key="nav_pricing", use_container_width=True):
                st.info("Pricing clicked")
            st.markdown('</div>', unsafe_allow_html=True)
            
        with m3:
            st.markdown('<div class="nav-link">', unsafe_allow_html=True)
            if st.button("Account", key="nav_account", use_container_width=True):
                st.info("Account clicked")
            st.markdown('</div>', unsafe_allow_html=True)

# 3. LANGUAGE BUTTONS (Top Right)
with col_nav_3:
    # Right align the buttons using a spacer
    c_space, c_btn_en, c_btn_fa = st.columns([2, 3, 3])
    
    with c_btn_en:
        btn_type = "primary" if st.session_state.lang == 'en' else "secondary"
        if st.button("EN", key="btn_en", type=btn_type, use_container_width=True):
            st.session_state.lang = 'en'; st.rerun()
            
    with c_btn_fa:
        btn_type = "primary" if st.session_state.lang == 'fa' else "secondary"
        if st.button("فارسی", key="btn_fa", type=btn_type, use_container_width=True):
            st.session_state.lang = 'fa'; st.rerun()

# Add a little spacing after the dark header before the content starts
st.markdown("<br><br>", unsafe_allow_html=True)

# D. HERO SECTION LOGIC
if st.session_state.df is None:
    render_hero_section()
    st.markdown("<h3 style='text-align: center; margin-top: 20px;'>👇 Get Started: Upload Your Data</h3>", unsafe_allow_html=True)    

# ==========================================
# 5. DATA SETTINGS EXPANDER (Hidden by default)
# ==========================================
# The "Filter" button that drops down options
with st.expander("📂 **Data Source & Settings**", expanded=(st.session_state.df is None)):
    c_upload, c_select = st.columns(2)
    
    with c_upload:
        # STEP 1: Enable multiple file selection
        uploaded_files = st.file_uploader(
            L('upload_csv'), 
            type=['csv', 'xlsx', 'xls'], 
            accept_multiple_files=True
        )
    
    with c_select:
        existing = [f for f in os.listdir(DATA_DIR) if f.endswith(('.csv', '.xlsx'))] if os.path.exists(DATA_DIR) else []
        selected_file = st.selectbox(L('or_select'), [''] + existing) if existing else None
    
    # Logic to load data - WITH FORMAT DETECTION
    if uploaded_files:
        all_dfs = []
        is_any_snappfood = False
        for file in uploaded_files:
            file_format = detect_file_format(file)
        
            if file_format == 'snappfood':
                is_any_snappfood = True
                temp_df = load_snappfood_file(file)
        
            else:
                temp_df = load_data(file)
                
            if temp_df is not None:
                all_dfs.append(temp_df)
            
        if all_dfs:
            # STEP 2: Combine all daily files into one master DataFrame
            df = pd.concat(all_dfs, ignore_index=True)
            st.session_state.is_snappfood = is_any_snappfood
            
            # STEP 3: Remove duplicate orders (safety check for overlapping files)
            if 'order_code' in df.columns:
                df = df.drop_duplicates(subset=['order_code'])
            
            # STEP 4: Global Keyword Filter (Pulling from config.py)
            from config import EXCLUDE_BRANCHES 
            if not df.empty and COLS['BRANCH'] in df.columns:
                pattern = '|'.join(EXCLUDE_BRANCHES)
                mask = df[COLS['BRANCH']].str.contains(pattern, case=False, na=False)
                df = df[~mask]
            
            # STEP 5: Finalize Session State
            st.session_state.df = df
            st.session_state.analyzer = ShilaAnalyzer(df, COLS)
            st.success(f"✅ Loaded {len(uploaded_files)} files! Total rows: {len(df):,}")
            
    elif selected_file:
        file_path = os.path.join(DATA_DIR, selected_file)
        file_format = detect_file_format(file_path)
        
        if file_format == 'snappfood':
            st.success("✅ SnappFood format detected")
            st.session_state.is_snappfood = True
            df = load_snappfood_file(file_path)
        else:
            st.success("✅ Original format detected")
            st.session_state.is_snappfood = False
            df = load_data(file_path)
        
        if df is not None:
            # Pull keywords from config
            from config import EXCLUDE_BRANCHES 
            if not df.empty and COLS['BRANCH'] in df.columns:
                # Join keywords into a regex pattern (e.g., "کیک خونه|کشمون|")
                pattern = '|'.join(EXCLUDE_BRANCHES)
                mask = df[COLS['BRANCH']].str.contains(pattern, case=False, na=False)
                df = df[~mask]
                
            st.session_state.df = df
            st.session_state.analyzer = ShilaAnalyzer(df, COLS)

if st.session_state.analyzer is None:
    st.info("👋 Please upload data or select a file from the settings menu above to begin.")
    st.stop()

# ==========================================
# 6. METRIC CARDS
# ==========================================
analyzer = st.session_state.analyzer
kpis = analyzer.get_kpis()

st.markdown(f"### {L('kpi_section')}")
st.markdown("<div style='margin-bottom: 20px'></div>", unsafe_allow_html=True) # Spacing

# Base metrics available for both formats
metrics_data = [
    (L('avg_rating'), f"{kpis['avg_rating']} / 5", "⭐", "#FFB020"),
    (L('total_orders'), f"{kpis['total_orders']:,}", "🛍️", "#2196F3"),
    (L('response_rate'), f"{kpis['response_rate']}%", "💬", "#9C27B0"),
]

# Only show NPS metrics if it's NOT a SnappFood file
if not st.session_state.get('is_snappfood', False):
    metrics_data.insert(0, (L('nps_score'), kpis['nps_score'], "📊", "#4CAF50" if kpis['nps_score'] > 0 else "#D32F2F"))
    metrics_data.append((L('promoters'), f"{kpis['promoters']:,}", "😊", "#4CAF50"))
    metrics_data.append((L('detractors'), f"{kpis['detractors']:,}", "😠", "#D32F2F"))

# Render Grid
cols = st.columns(len(metrics_data)) # Dynamic column count
for i, col in enumerate(cols):
    label, value, icon, color = metrics_data[i]
    col.markdown(get_metric_html(label, value, icon, color), unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# 7. MAIN TABS (Complete Content)
# ==========================================
# --- TAB DEFINITIONS ---
# Define base tabs available for everyone
main_tabs = [L('tab_overview')]

# Only add NPS/Tag dependent tabs if it's NOT SnappFood
if not st.session_state.get('is_snappfood', False):
    main_tabs.append(L('tab_pareto'))
    main_tabs.append(L('tab_kano'))

main_tabs.extend([L('tab_branches'), L('tab_aspects'), L('tab_trends'), L('tab_ai')])
main_tabs.extend(["🍔 Products", "📝 Text Mining", "🤖 Machine Learning"])

# Create the tabs object
tabs = st.tabs(main_tabs)

# Define a counter to track the current index
t = 0

# Helper to clean up chart look
def clean_chart(fig, height=400):
    fig.update_layout(
        height=height,
        margin=dict(t=30, b=40, l=40, r=40),
        paper_bgcolor='rgba(0,0,0,0)',   # Transparent background
        plot_bgcolor='rgba(0,0,0,0)',    # Transparent plot area
        font=dict(family="Inter, sans-serif", color="#3C4257"),
        xaxis=dict(showgrid=False, color="#697386"),
        yaxis=dict(showgrid=True, gridcolor="#E1E4E8", color="#697386") # Light grey grid lines
    )
    return fig

# TAB 1: OVERVIEW
with tabs[t]:
    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    
    with c1:
        st.markdown(f"### {L('rating_distribution')}")
        rd = analyzer.get_rating_distribution()
        if not rd.empty:
            fig = px.bar(rd, x='Rating', y='Count', color='Rating', 
                         color_continuous_scale=['#D32F2F','#FF9800','#FFEB3B','#8BC34A','#4CAF50'], 
                         text='Count')
            fig.update_traces(textposition='outside')
            st.plotly_chart(clean_chart(fig, 350), use_container_width=True)
            
    with c2:
    # Only show NPS chart if NOT SnappFood
        if not st.session_state.get('is_snappfood', False):
            st.markdown(f"### {L('nps_distribution')}")
            nd = analyzer.get_nps_distribution()
            if not nd.empty:
                fig = px.bar(
                    nd, x='NPS', y='Count', color='Segment',
                    color_discrete_map={'Promoter':'#4CAF50','Passive':'#FF9800','Detractor':'#D32F2F'},
                    text='Count'
                )
                st.plotly_chart(clean_chart(fig, 350), use_container_width=True)
        else:
            # OPTIONAL: Show something else for SnappFood, like a 5-star summary
            st.markdown("### 📊 Rating Summary")
            st.info("SnappFood files provide Star Ratings (1-5) instead of NPS (0-10). View the Rating Distribution chart on the left for details.")
    
    # Tables with clean styling
    # Hide the "Top Issues" and "Top Strengths" tables for SnappFood too
    if not st.session_state.get('is_snappfood', False):
        st.markdown("---")
        c3, c4 = st.columns(2)
        with c3:
            st.markdown(f"#### 🚨 {L('top_issues')}")
            st.dataframe(analyzer.get_top_issues(10), use_container_width=True, hide_index=True)
        with c4:
            st.markdown(f"#### 🏆 {L('top_strengths')}")
            st.dataframe(analyzer.get_top_strengths(10), use_container_width=True, hide_index=True)
    
    # ==========================================
    # CUSTOMER SEGMENTS SECTION (ENHANCED)
    # ==========================================
    if not st.session_state.get('is_snappfood', False):
        st.markdown("---")
        st.markdown("#### 🎯 Customer Segments")
        
        recovery = analyzer.get_recovery_opportunities()
        
        if not recovery.empty:
            # A. SEGMENT METRIC CARDS
            cols_seg = st.columns(len(recovery))
            segment_colors = {
                'Happy': '#4CAF50', 'At Risk': '#D32F2F', 
                'Silent Churner': '#FF9800', 'Recovery': '#2196F3', 'Neutral': '#9E9E9E'
            }            
            for i, (_, row) in enumerate(recovery.iterrows()):
                with cols_seg[i]:
                    color = segment_colors.get(row['segment'], '#9E9E9E')
                    st.markdown(f"""
                    <div style="background-color:{color}; padding:12px; border-radius:8px; text-align:center;">
                        <span style="font-size:24px;">{row['emoji']}</span><br>
                        <b style="color:white; font-size:14px;">{row['segment_fa'] if st.session_state.lang == 'fa' else row['segment']}</b><br>
                        <span style="color:white; font-size:22px;"><b>{row['count']:,}</b></span><br>
                        <small style="color:white;">({row['percentage']}%)</small>
                    </div>
                    """, unsafe_allow_html=True)
    
            st.markdown("<br>", unsafe_allow_html=True)
        
        # ==========================================
        # SEGMENTATION MATRIX
        # ==========================================
        with st.expander("📊 View Segmentation Logic Matrix", expanded=False):
            
            st.caption("How customers are categorized based on Rating + NPS combination")
            
            col_matrix, col_legend = st.columns([3, 2])
            
            with col_matrix:
                # Create the matrix heatmap
                fig_matrix = go.Figure()
                
                fig_matrix.add_trace(go.Heatmap(
                    z=[
                        [1, 4, 4],   # Low Rating (1-2)
                        [2, 3, 4],   # Mid Rating (3)
                        [2, 3, 5]    # High Rating (4-5)
                    ],
                    x=['Low NPS<br>(0-6)', 'Mid NPS<br>(7-8)', 'High NPS<br>(9-10)'],
                    y=['Low Rating (1-2)', 'Mid Rating (3)', 'High Rating (4-5)'],
                    text=[
                        ['🚨<br><b>At Risk</b><br>در خطر', '🔄<br><b>Recovery</b><br>قابل بازیابی', '🔄<br><b>Recovery</b><br>قابل بازیابی'],
                        ['⚠️<br><b>Silent Churner</b><br>ریزش خاموش', '😐<br><b>Neutral</b><br>خنثی', '🔄<br><b>Recovery</b><br>قابل بازیابی'],
                        ['⚠️<br><b>Silent Churner</b><br>ریزش خاموش', '😐<br><b>Neutral</b><br>خنثی', '😊<br><b>Happy</b><br>راضی']
                    ],
                    texttemplate='%{text}',
                    textfont=dict(size=11, color='white'),
                    colorscale=[
                        [0.0, '#D32F2F'],   # At Risk - Red
                        [0.25, '#FF9800'],  # Silent Churner - Orange
                        [0.5, '#9E9E9E'],   # Neutral - Gray
                        [0.75, '#2196F3'],  # Recovery - Blue
                        [1.0, '#4CAF50']    # Happy - Green
                    ],
                    showscale=False,
                    hoverinfo='skip'
                ))
                
                fig_matrix.update_layout(
                    height=280,
                    margin=dict(l=10, r=10, t=30, b=10),
                    xaxis=dict(
                        title="NPS Score →",
                        side='top',
                        tickfont=dict(size=11)
                    ),
                    yaxis=dict(
                        title="",
                        tickfont=dict(size=11)
                    ),
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(0,0,0,0)'
                )
                
                st.plotly_chart(fig_matrix, use_container_width=True)
            
            with col_legend:
                st.markdown("**Segment Definitions:**")
                
                st.markdown("""
                <div style="background-color:#4CAF50; padding:6px 10px; border-radius:5px; margin:4px 0;">
                    <span style="color:white;">😊 <b>Happy</b> - Rating ≥4 & NPS ≥9</span>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("""
                <div style="background-color:#D32F2F; padding:6px 10px; border-radius:5px; margin:4px 0;">
                    <span style="color:white;">🚨 <b>At Risk</b> - Rating ≤2 & NPS ≤6</span>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("""
                <div style="background-color:#FF9800; padding:6px 10px; border-radius:5px; margin:4px 0;">
                    <span style="color:white;">⚠️ <b>Silent Churner</b> - Rating ≥3 & NPS ≤6</span>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("""
                <div style="background-color:#2196F3; padding:6px 10px; border-radius:5px; margin:4px 0;">
                    <span style="color:white;">🔄 <b>Recovery</b> - Rating ≤3 & NPS ≥7</span>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("""
                <div style="background-color:#9E9E9E; padding:6px 10px; border-radius:5px; margin:4px 0;">
                    <span style="color:white;">😐 <b>Neutral</b> - Everyone else</span>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Action recommendations (compact)
            st.markdown("**🎯 Quick Actions:**")
            
            col_act1, col_act2, col_act3, col_act4 = st.columns(4)
            
            with col_act1:
                st.markdown("""
                <div style="border-left: 3px solid #D32F2F; padding-left: 8px; font-size:12px;">
                    <b>🚨 At Risk</b><br>
                    Urgent recovery!<br>
                    Apology + discount
                </div>
                """, unsafe_allow_html=True)
            
            with col_act2:
                st.markdown("""
                <div style="border-left: 3px solid #FF9800; padding-left: 8px; font-size:12px;">
                    <b>⚠️ Silent Churner</b><br>
                    Survey to understand<br>
                    Re-engage campaign
                </div>
                """, unsafe_allow_html=True)
            
            with col_act3:
                st.markdown("""
                <div style="border-left: 3px solid #2196F3; padding-left: 8px; font-size:12px;">
                    <b>🔄 Recovery</b><br>
                    Easy win!<br>
                    One good experience
                </div>
                """, unsafe_allow_html=True)
            
            with col_act4:
                st.markdown("""
                <div style="border-left: 3px solid #4CAF50; padding-left: 8px; font-size:12px;">
                    <b>😊 Happy</b><br>
                    Retain & reward<br>
                    Ask for referrals
                </div>
                """, unsafe_allow_html=True)

# Increment tab counter for next tab
t += 1

# TAB 2: PARETO ANALYSIS
if not st.session_state.get('is_snappfood', False):
    with tabs[t]:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"### {L('issues_by_damage')}")
        st.caption("Which issues are hurting your star rating the most?")
    
        pareto = analyzer.get_pareto_analysis()
        if len(pareto) > 0:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
        
            # Bar Chart (Damage)
            fig.add_trace(go.Bar(
                x=pareto['tag'].head(15), 
                y=pareto['total_damage'].head(15), 
                name='Impact Score', 
                marker_color='#D32F2F', 
                opacity=0.85
            ), secondary_y=False)
        
            # Line Chart (Cumulative %)
            fig.add_trace(go.Scatter(
                x=pareto['tag'].head(15), 
                y=pareto['cumulative_pct'].head(15), 
                name='Cumulative %', 
                mode='lines+markers', 
                line=dict(color='#1A1F36', width=2),
                marker=dict(size=6)
            ), secondary_y=True)
        
            # The 80% Rule Line
            fig.add_hline(y=80, line_dash="dash", line_color="#4CAF50", secondary_y=True, annotation_text="80% Cutoff")
        
            fig = clean_chart(fig, 500)
            fig.update_layout(xaxis_tickangle=-45, legend=dict(orientation="h", y=1.1))
            st.plotly_chart(fig, use_container_width=True)
            
            with st.expander("View Detailed Data Table"):
                st.dataframe(pareto.head(20), use_container_width=True, hide_index=True)
        else:
            st.info("Not enough issue data to generate Pareto analysis.")
            
    # CRITICAL: Increment the counter only inside the IF block 
    # because the tab only exists in the list for standard files.
    t += 1

# TAB: KANO MODEL
if not st.session_state.get('is_snappfood', False):
    with tabs[t]:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"### {L('kano_classification')}")
        st.caption("Classifying features into Must-Be, Performance, and Delighters based on customer feedback.")
    
        kano = analyzer.get_kano_analysis()
        if len(kano) > 0:
            fig = px.scatter(
                kano, 
                x='lift_as_strength', 
                y='drop_as_weakness', 
                color='kano_type', 
                hover_name='attribute', 
                size='strength_mentions', 
                size_max=40,
                color_discrete_map={'Must-Be':'#D32F2F', 'Performance':'#FFB020', 'Delighter':'#4CAF50'}
            )
        
            # Quadrant Lines
            fig.add_hline(y=0.5, line_dash="dot", line_color="#E1E4E8")
            fig.add_vline(x=0.3, line_dash="dot", line_color="#E1E4E8")
            
            # Labels for Quadrants
            fig.add_annotation(x=0.8, y=0.1, text="Delighters (Unique)", showarrow=False, font=dict(color="green"))
            fig.add_annotation(x=0.1, y=0.9, text="Must-Be (Critical)", showarrow=False, font=dict(color="red"))
            
            st.plotly_chart(clean_chart(fig, 500), use_container_width=True)

        else:
            st.info("Not enough strength/weakness data to generate Kano classification.")

    # Increment counter only if we actually processed this tab
    t += 1

# TAB: BRANCH COMPARISON
with tabs[t]:
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"### {L('branch_comparison')}")
    
    br_stats, br_issues = analyzer.get_branch_analysis()
    
    if not br_stats.empty:
        # Dynamic average for the chart line
        avg_all = br_stats['avg_rating'].mean()
        
        # ==========================================
        # DIVERGING BAR CHART - Rating vs Average
        # ==========================================
        st.markdown("#### 📊 Branch Performance vs Company Average")
        
        # Sort by rating_vs_avg for better visualization
        br_sorted = br_stats.sort_values('rating_vs_avg', ascending=True)
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=br_sorted['branch'],
            x=br_sorted['rating_vs_avg'],
            orientation='h',
            marker=dict(
                color=br_sorted['rating_vs_avg'],
                colorscale=[
                    [0.0, '#D32F2F'],    # Red for worst
                    [0.25, '#FF9800'],   # Orange
                    [0.5, '#FFEB3B'],    # Yellow for average
                    [0.75, '#8BC34A'],   # Light green
                    [1.0, '#4CAF50']     # Green for best
                ],
                cmid=0,
                colorbar=dict(
                    title="vs Avg",
                    tickvals=[-0.5, -0.25, 0, 0.25, 0.5],
                    ticktext=['🔴 -0.5', '🟠 -0.25', '🟡 0', '🟢 +0.25', '🏆 +0.5']
                )
            ),
            text=[f"{x:+.2f}" for x in br_sorted['rating_vs_avg']],
            textposition='outside',
            textfont=dict(size=10),
            hovertemplate=(
                "<b>%{y}</b><br>" +
                "Rating: " + br_sorted['avg_rating'].round(2).astype(str) + "<br>" +
                "vs Average: %{x:+.2f}<br>" +
                "<extra></extra>"
            )
        ))
        
        # Add vertical line at 0 (average)
        fig.add_vline(x=0, line_dash="solid", line_color="#1A1F36", line_width=2)
        
        # Annotations
        fig.add_annotation(
            x=0, y=1.02, yref="paper",
            text="📍 Company Average (3.93)",
            showarrow=False,
            font=dict(size=11, color="#1A1F36")
        )
        
        fig.update_layout(
            height=max(450, len(br_sorted) * 22),
            margin=dict(l=10, r=80, t=40, b=40),
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(
                title="Rating vs Company Average",
                zeroline=True,
                zerolinecolor='#1A1F36',
                zerolinewidth=2,
                gridcolor='#E1E4E8',
                range=[min(br_sorted['rating_vs_avg'].min() - 0.15, -0.5), 
                       max(br_sorted['rating_vs_avg'].max() + 0.15, 0.5)]
            ),
            yaxis=dict(
                title="",
                tickfont=dict(size=10),
                automargin=True
            ),
            showlegend=False
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        # ==========================================
        # COLOR SCALE LEGEND
        # ==========================================
        st.markdown("""
        <div style="background: linear-gradient(90deg, #D32F2F 0%, #FF9800 25%, #FFEB3B 50%, #8BC34A 75%, #4CAF50 100%); 
                    height: 15px; border-radius: 8px; margin: 5px 0;"></div>
        """, unsafe_allow_html=True)
        
        col_legend = st.columns(5)
        with col_legend[0]:
            st.markdown("<center>🔴 **Poor**<br><small>< -0.3</small></center>", unsafe_allow_html=True)
        with col_legend[1]:
            st.markdown("<center>🟠 **Weak**<br><small>-0.3 to 0</small></center>", unsafe_allow_html=True)
        with col_legend[2]:
            st.markdown("<center>🟡 **Average**<br><small>≈ 0</small></center>", unsafe_allow_html=True)
        with col_legend[3]:
            st.markdown("<center>🟢 **Good**<br><small>0 to +0.3</small></center>", unsafe_allow_html=True)
        with col_legend[4]:
            st.markdown("<center>🏆 **Great**<br><small>> +0.3</small></center>", unsafe_allow_html=True)
        
        st.markdown("---")
        
        # ==========================================
        # SUMMARY METRICS
        # ==========================================
        above_avg = len(br_stats[br_stats['rating_vs_avg'] > 0])
        below_avg = len(br_stats[br_stats['rating_vs_avg'] < 0])
        great_branches = len(br_stats[br_stats['rating_vs_avg'] > 0.3])
        poor_branches = len(br_stats[br_stats['rating_vs_avg'] < -0.3])
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
            st.metric("🟢 Above Average", f"{above_avg} branches", f"{above_avg/len(br_stats)*100:.0f}%")
        with col_m2:
            st.metric("🔴 Below Average", f"{below_avg} branches", f"{below_avg/len(br_stats)*100:.0f}%")
        with col_m3:
            st.metric("🏆 Great Performers", f"{great_branches} branches")
        with col_m4:
            st.metric("⚠️ Need Attention", f"{poor_branches} branches")
        
        st.markdown("---")
        
        # ==========================================
        # TOP & BOTTOM BRANCHES
        # ==========================================
        col_top, col_bottom = st.columns(2)
        
        with col_top:
            st.markdown("#### 🏆 Top 5 Branches")
            top5 = br_stats.head(5)[['branch', 'avg_rating', 'rating_vs_avg', 'nps_score', 'order_count']].copy()
            top5['status'] = top5['rating_vs_avg'].apply(lambda x: '🏆' if x > 0.3 else '🟢')
            st.dataframe(top5, use_container_width=True, hide_index=True)
        
        with col_bottom:
            st.markdown("#### ⚠️ Bottom 5 Branches")
            bottom5 = br_stats.tail(5)[['branch', 'avg_rating', 'rating_vs_avg', 'nps_score', 'order_count']].copy()
            bottom5['status'] = bottom5['rating_vs_avg'].apply(lambda x: '🔴' if x < -0.3 else '🟠')
            st.dataframe(bottom5, use_container_width=True, hide_index=True)
        
        # ==========================================
        # FULL DATA TABLE (Expandable)
        # ==========================================
        with st.expander("📋 View All Branch Data"):
            br_display = br_stats.copy()
            br_display['status'] = br_display['rating_vs_avg'].apply(
                lambda x: '🏆 Great' if x > 0.3 else ('🟢 Good' if x > 0 else ('🟠 Weak' if x > -0.3 else '🔴 Poor'))
            )
            st.dataframe(br_display, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # ==========================================
    # BRANCH × PRODUCT HEATMAP (Keep as before)
    # ==========================================
    st.markdown("#### 🔥 Branch × Product Performance")
    matrix = analyzer.get_branch_product_matrix()
    if not matrix.empty:
        # 1. Handle large matrices: If there are > 15 products, make the chart wider
        height_calc = max(400, len(matrix.index) * 30) 
    
        fig_heat = px.imshow(
            matrix,
            # Red for bad, Yellow for mid, Green for great
            color_continuous_scale=['#D32F2F', '#FFEB3B', '#4CAF50'],
            aspect='auto',
            text_auto='.1f', # Single decimal is cleaner for heatmaps
            labels=dict(color="Avg Rating")
        )
    
        fig_heat.update_layout(
            height=height_calc,
            xaxis_title="Product",
            yaxis_title="Branch",
            margin=dict(l=10, r=10, t=30, b=10)
        )
        
        # Use container width but allow the height to scale
        st.plotly_chart(fig_heat, use_container_width=True)
    else:
        st.info("No product data available for this selection. Check if 'نام محصولات سفارشی' is in your file.")

t += 1

## TAB: ASPECT SENTIMENT
with tabs[t]:
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"### {L('aspect_sentiment')}")
    
    aspects = analyzer.get_aspect_sentiment()
    if len(aspects) > 0:
        fig = px.bar(
            aspects, 
            y='aspect', 
            x='sentiment_score', 
            orientation='h', 
            color='sentiment_score', 
            color_continuous_scale=['#D32F2F', '#FFEB3B', '#4CAF50'], 
            color_continuous_midpoint=0
        )
        fig.add_vline(x=0, line_color="#3C4257", line_width=1)
        fig.update_layout(yaxis={'categoryorder':'total ascending'}) # Sort bars
        
        st.plotly_chart(clean_chart(fig, 500), use_container_width=True)
    else:
        # Fallback if no comments or keywords were found
        st.info("💡 Not enough comment data found to perform aspect-based sentiment analysis.")

# Increment tab counter for next tab
t += 1

# TAB: TRENDS
with tabs[t]:
    st.markdown("<br>", unsafe_allow_html=True)
    is_sf = st.session_state.get('is_snappfood', False)
    low_comments = pd.DataFrame()
    peaks = None
    
    if is_sf:
        st.markdown("### 🚀 Peak Hour Insights")
        peaks = analyzer.get_peak_hour_analysis()
        
        if peaks:
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("🔥 Busiest Hour", 
                          f"{int(peaks['busiest_hour']):02d}:00", 
                          f"{int(peaks['peak_volume'])} orders")
            with c2:
                st.metric("⭐ Best Service", 
                          f"{int(peaks['best_hour']):02d}:00", 
                          f"{peaks['best_rating']:.2f} Rating")
            with c3:
                # Color this one red if the rating is low
                st.metric("⚠️ Pressure Point", 
                          f"{int(peaks['worst_hour']):02d}:00", 
                          f"{peaks['worst_rating']:.2f} Rating",
                          delta_color="inverse")
            st.markdown("---")
            
    # 1. DYNAMIC LOGIC: Hourly for SnappFood, Daily for standard
    # ==========================================
    # TIME GRANULARITY SELECTOR
    # ==========================================
    if is_sf:
        st.markdown("### 📈 Performance Trends")
        granularity = st.radio(
            "Select time view:",
            ["🕒 Hourly", "📅 Daily", "📆 Weekly", "🗓️ Monthly"],
            horizontal=True,
            key="trend_granularity"
        )
    else:
        granularity = "📅 Daily"  # Default for standard files

    # ==========================================
    # LOAD DATA BASED ON GRANULARITY
    # ==========================================
    if granularity == "🕒 Hourly":
        st.markdown(f"### 🕒 {L('hourly_rating_trend')}")
        st.caption("Performance across the 24-hour cycle")
        data = analyzer.get_hourly_trends()
        x_col = 'hour'
        x_label = "Hour (00:00 - 23:00)"

    elif granularity == "📅 Daily":
        st.markdown(f"### 📅 {L('daily_trend')}")
        st.caption("Day-by-day performance")
        data = analyzer.get_daily_trends()
        x_col = 'date'
        x_label = "Date"

    elif granularity == "📆 Weekly":
        st.markdown("### 📆 Weekly Trends")
        st.caption("Aggregated weekly performance")
        data = analyzer.get_weekly_trends()  # NEW METHOD NEEDED
        st.write("DEBUG weekly:", len(data), "rows")  # ← ADD
        st.dataframe(data)                             # ← ADD
        x_col = 'week'
        x_label = "Week"

    elif granularity == "🗓️ Monthly":
        st.markdown("### 🗓️ Monthly Trends")
        st.caption("Month-over-month performance")
        data = analyzer.get_monthly_trends()  # NEW METHOD NEEDED
        x_col = 'month'
        x_label = "Month"

    # ==========================================
    # RENDER CHART (works for all granularities)
    # ==========================================
    if not data.empty:
        fig = make_subplots(
            rows=2, cols=1, 
            subplot_titles=('Average Rating', 'Order Volume'), 
            vertical_spacing=0.2
        )
        
        # Row 1: Rating Trend
        fig.add_trace(go.Scatter(
            x=data[x_col], y=data['avg_rating'], 
            mode='lines+markers', name='Rating', 
            line=dict(color='#2196F3', width=3)
        ), row=1, col=1)

        # Add rolling average line where available
        if 'rating_7day_avg' in data.columns:
            fig.add_trace(go.Scatter(
                x=data[x_col], y=data['rating_7day_avg'], 
                mode='lines', name='7-Day Trend', 
                line=dict(color='#E1E4E8', width=2)
            ), row=1, col=1)
        elif 'rating_4week_avg' in data.columns:
            fig.add_trace(go.Scatter(
                x=data[x_col], y=data['rating_4week_avg'], 
                mode='lines', name='4-Week Trend', 
                line=dict(color='#E1E4E8', width=2)
            ), row=1, col=1)
        
        # Row 2: Order Volume
        fig.add_trace(go.Bar(
            x=data[x_col], y=data['order_count'], 
            name='Orders', marker_color='#4CAF50', opacity=0.7
        ), row=2, col=1)

        # Axis Formatting
        fig.update_xaxes(title_text=x_label, row=2, col=1)
        
        # Force integer ticks for hourly view
        if granularity == "🕒 Hourly":
            fig.update_xaxes(tickmode='linear', tick0=0, dtick=1, row=1, col=1)
            fig.update_xaxes(tickmode='linear', tick0=0, dtick=1, row=2, col=1)
        
        # Rotate labels for weekly/monthly if many data points
        if granularity in ["📆 Weekly", "🗓️ Monthly"]:
            fig.update_xaxes(tickangle=-45, row=1, col=1)
            fig.update_xaxes(tickangle=-45, row=2, col=1)

        st.plotly_chart(clean_chart(fig, 600), use_container_width=True)
        
        # ==========================================
        # CHANGE METRICS (for weekly/monthly)
        # ==========================================
        if granularity in ["📆 Weekly", "🗓️ Monthly"] and len(data) > 1:
            st.markdown("---")
            latest = data.iloc[-1]
            prev = data.iloc[-2]
            
            period_label = "Week" if granularity == "📆 Weekly" else "Month"
            
            cols_metric = st.columns(3)
            cols_metric[0].metric(
                f"Latest {period_label}", 
                latest[x_col],
                f"Rating: {latest['avg_rating']}"
            )
            
            rating_delta = latest['avg_rating'] - prev['avg_rating']
            cols_metric[1].metric(
                "Rating Change", 
                f"{latest['avg_rating']:.2f}", 
                f"{rating_delta:+.2f}"
            )
            
            if prev['order_count'] > 0:
                order_delta_pct = ((latest['order_count'] - prev['order_count']) / prev['order_count']) * 100
            else:
                order_delta_pct = 0
            cols_metric[2].metric(
                "Orders Change", 
                f"{int(latest['order_count']):,}", 
                f"{order_delta_pct:+.1f}%"
            )
    else:
        st.info("Insufficient data to generate trends for this time period.")
    
    # 2. MONTH-OVER-MONTH (Hide for single-day SnappFood files)
    if not is_sf:
        st.markdown("---")
        st.markdown(f"#### 📅 {L('mom_comparison')}")
        mom = analyzer.get_mom_comparison()
        
        if not mom.empty:
            fig_mom = make_subplots(specs=[[{"secondary_y": True}]])
            fig_mom.add_trace(go.Bar(x=mom['year_month'], y=mom['order_count'], name='Orders', marker_color='#4CAF50', opacity=0.6), secondary_y=False)
            fig_mom.add_trace(go.Scatter(x=mom['year_month'], y=mom['avg_rating'], name='Avg Rating', mode='lines+markers', line=dict(color='#2196F3', width=3)), secondary_y=True)
            st.plotly_chart(clean_chart(fig_mom, 400), use_container_width=True)

            # Change Indicators
            if len(mom) > 1:
                latest = mom.iloc[-1]
                cols_metric = st.columns(3) # Always 3 columns, we'll leave one empty if needed
                cols_metric[0].metric("Rating Change", f"{latest['avg_rating']:.2f}", f"{latest['rating_change']:+.2f}")
                cols_metric[1].metric("Orders Change", f"{latest['order_count']:,}", f"{latest['orders_change_pct']:+.1f}%")
                
                # NPS only for standard
                if 'nps_score' in latest and 'nps_change' in latest:
                    cols_metric[2].metric("NPS Change", f"{latest['nps_score']:.1f}", f"{latest['nps_change']:+.1f}")
    
    if is_sf:
        st.markdown("---")
        st.markdown("## 🕵️ Weekly Low-Rating Deep Dive (1-3 Stars)")

        # Get data from both analysis methods
        low_comments = analyzer.get_low_rating_comments_by_hour()
        topic_summary, weekly_trend = analyzer.get_low_rating_deep_dive()

        if not weekly_trend.empty or not topic_summary.empty:
            # 1. THE BIG PICTURE (Weekly Trends & Topics)
            col1, col2 = st.columns([1, 1])
            with col1:
                st.write("📅 **Weekly Trend of Complaints**")
                fig_trend = px.line(weekly_trend, analyzer.cols.get('CREATED_AT'), y='complaint_count', # ensure x matches your method's output
                                    title="Total Low Ratings per Day",
                                    line_shape='spline')
                fig_trend.update_traces(line_color='#d32f2f', fill='tozeroy')
                st.plotly_chart(fig_trend, use_container_width=True)

            with col2:
                st.write("📊 **Main Topics causing Low Ratings**")
                overall_topics = topic_summary.groupby('topics')['count'].sum().sort_values()
                fig_topics = px.bar(overall_topics, orientation='h', color_discrete_sequence=['#d32f2f'])
                st.plotly_chart(fig_topics, use_container_width=True)

            # 2. THE BRANCH HEATMAP
            st.write("🏪 **Problem Heatmap: Which Branches have which Problems?**")
            topic_matrix = topic_summary.pivot(index=analyzer.cols.get('BRANCH'), 
                                               columns='topics', values='count').fillna(0)
            fig_heatmap = px.imshow(topic_matrix, text_auto=True, color_continuous_scale='Reds', aspect='auto')
            st.plotly_chart(fig_heatmap, use_container_width=True)
            
            # 3. THE DRILL DOWN (Interactive Comment Reader)
            st.markdown("---")
            st.write("💬 **Drill Down: Read specific complaints by hour**")
        
            selected_hour = st.select_slider(
            "Select an hour to read comments:",
            options=sorted(low_comments['hour'].unique()),
            format_func=lambda x: f"{int(x):02d}:00"
            )
        
            hour_filtered = low_comments[low_comments['hour'] == selected_hour]
            for _, row in hour_filtered.iterrows():
                with st.chat_message("user", avatar="🚨"):
                    st.write(f"**Rating: {int(row[analyzer.cols.get('RATING')])}** | Branch: {row[analyzer.cols.get('BRANCH')]}")
                    st.info(row[analyzer.cols.get('COMMENT')])
        else:
            st.success("🌟 Incredible! No 1-3 star ratings found in this week's data.")

        st.markdown("---")
        st.markdown("### 📋 Deep Dive: What's inside 'Other' & 'Uncategorized'?")

        col_select, _ = st.columns([1, 2])
        with col_select:
            view_type = st.selectbox("Select category to inspect:", ["Other", "Uncategorized"])

        unmapped_data = analyzer.get_unmapped_comments(category_type=view_type)

        if not unmapped_data.empty:
            st.write(f"Found **{len(unmapped_data)}** comments in **{view_type}**.")
    
        # Display as a searchable table
            st.dataframe(
                unmapped_data, 
                use_container_width=True,
                column_config={
                    analyzer.cols.get('COMMENT'): st.column_config.TextColumn("Raw Feedback", width="large"),
                    analyzer.cols.get('RATING'): st.column_config.NumberColumn("Rating", format="%d ⭐")
                }
            )
    
        # Pro-tip for the user
        if view_type == "Other":
            st.info("💡 **Tip:** Look for recurring words in these comments. Add them to your `ASPECTS` dictionary in `config.py` to move these rows into a specific category!")
        else:
            st.success(f"No data found for {view_type}.")
        
# Increment main tab counter
t += 1

# TAB: AI INSIGHTS
with tabs[t]:
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(f"## 🤖 {L('ai_title')}")
    
    # Initialize Generator
    gen = InsightsGenerator(lang=st.session_state.lang)
    summary = analyzer.get_summary_for_ai()
    
    if st.session_state.get('is_snappfood', False):
        summary['data_format'] = 'snappfood_star_rating'
        
    c_ai_1, c_ai_2 = st.columns([2, 1])
    
    with c_ai_1:
        st.markdown(f"### {L('ai_rule_based')}")
        # Display automated insights in a nice list
        rule_insights = gen.generate_rule_based_insights(summary)
        if rule_insights:
            for insight in rule_insights:
                if insight:
                    st.info(insight, icon="💡")
        else:
            st.write("No automated insights available for this dataset.")
    
    with c_ai_2:
        st.markdown(f"### {L('ai_claude')}")
        
        # Check API Key
        from config import ANTHROPIC_API_KEY
        if not ANTHROPIC_API_KEY:
            st.warning(L('api_key_missing'))
            with st.expander("Setup Instructions"):
                st.markdown(get_api_setup_instructions())
        else:
            # Chat Interface
            q = st.text_area("Ask AI about your data:", placeholder="e.g., How can we improve delivery speed?", height=100)
            
            if st.button(L('generate_insights'), type="primary", use_container_width=True):
                with st.spinner("Analyzing data..."):
                    result = gen.generate_claude_insights(summary, q)
                    if result['success']:
                        st.markdown("### AI Response:")
                        st.markdown(result['insights'])
                    else:
                        st.error(result.get('error'))

t += 1

# TAB: PRODUCTS
with tabs[t]:  # Index 1 for Products
    st.markdown("<br>", unsafe_allow_html=True)
    
    c1, c2 = st.columns(2)
    
    with c1:
        st.markdown("### 🍔 Product Performance")
        products = analyzer.get_product_analysis()
        if len(products) > 0:
            fig = px.bar(
                products.head(15), 
                x='product', y='avg_rating',
                color='avg_rating',
                color_continuous_scale=['#D32F2F', '#FFB020', '#4CAF50'],
                text='avg_rating'
            )
            fig.update_traces(texttemplate='%{text:.2f}', textposition='outside')
            st.plotly_chart(clean_chart(fig, 400), use_container_width=True)
        else:
            st.info("No product data available")
    
    with c2:
        if not st.session_state.get('is_snappfood', False):
            st.markdown("### ⚠️ Issue Category Impact")
            issue_cats = analyzer.get_issue_category_analysis()
            if len(issue_cats) > 0:
                fig = px.bar(
                    issue_cats,
                    x='category_fa' if st.session_state.lang == 'fa' else 'category',
                    y='rating_impact',
                    color='rating_impact',
                    color_continuous_scale=['#4CAF50', '#FFB020', '#D32F2F'],
                    text='issue_count'
                )
                fig.update_traces(texttemplate='%{text} issues', textposition='outside')
                fig.update_yaxes(title="Rating Impact (higher = worse)")
                st.plotly_chart(clean_chart(fig, 400), use_container_width=True)
            
                # Details table
                st.dataframe(issue_cats, use_container_width=True, hide_index=True)
            else:
                st.info("No issue category data available")
        else:
            # For SnappFood, provide a useful alternative or an explanation
            st.markdown("### 📋 Analysis Note")
            st.warning("""
                **Issue Category Analysis is unavailable for SnappFood files.**
                
                This analysis requires specific structured data (Delivery, Packaging, and Personnel columns) 
                found in the original feedback format. 
            """)
            st.info("Check the **Aspect Sentiment** tab for text-based insights into product and service quality.")

# Increment tab counter
t += 1

# TAB : TEXT MINING
with tabs[t]:  # Adjust index as needed
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### 📝 Text Mining Analysis")
    st.caption("Deep analysis of customer comments and feedback")
    
    # Check if text column exists
    text_col = analyzer.get_text_column()
    
    if text_col:
        
        # ==========================================
        # WORD CLOUD
        # ==========================================
        st.markdown("#### ☁️ Word Cloud - Most Frequent Words")

        word_freq = analyzer.get_word_frequency(min_freq=5, top_n=100)

        if word_freq:
            try:
                from wordcloud import WordCloud
                import arabic_reshaper
                from bidi.algorithm import get_display
                import matplotlib.pyplot as plt
                import os
                import platform
        
                # Reshape Persian text for correct display
                reshaped_freq = {}
                for word, freq in word_freq.items():
                    try:
                        reshaped = arabic_reshaper.reshape(word)
                        bidi_text = get_display(reshaped)
                        reshaped_freq[bidi_text] = freq
                    except:
                        reshaped_freq[word] = freq
        
                # Find appropriate font based on OS
                font_path = None
        
                if platform.system() == 'Windows':
                    # Windows fonts that support Persian/Arabic
                    possible_fonts = [
                        'C:/Windows/Fonts/Vazirmatn-Regular.ttf',
                        'C:/Windows/Fonts/tahoma.ttf',       # Tahoma (best for Persian)
                        'C:/Windows/Fonts/arial.ttf',        # Arial
                        'C:/Windows/Fonts/segoeui.ttf',      # Segoe UI
                        'C:/Windows/Fonts/times.ttf',        # Times New Roman
                        'C:/Windows/Fonts/calibri.ttf',      # Calibri
                    ]
                elif platform.system() == 'Darwin':  # macOS
                    possible_fonts = [
                        '/Library/Fonts/Arial.ttf',
                        '/System/Library/Fonts/Helvetica.ttc',
                    ]
                else:  # Linux
                    possible_fonts = [
                        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
                        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                    ]
        
                # Find first available font
                for font in possible_fonts:
                    if os.path.exists(font):
                        font_path = font
                        break
        
                if font_path is None:
                    # Use default (may not support Persian well)
                    st.warning("⚠️ No suitable font found. Persian text may not display correctly.")
                    font_path = None
                    
                # Generate word cloud
                wc = WordCloud(
                    font_path=font_path,
                    width=800, 
                    height=400,
                    background_color='white',
                    colormap='viridis',
                    max_words=100,
                    prefer_horizontal=0.7,
                    min_font_size=10
                    ).generate_from_frequencies(reshaped_freq)
                
                fig_wc, ax = plt.subplots(figsize=(12, 6))
                ax.imshow(wc, interpolation='bilinear')
                ax.axis('off')
                st.pyplot(fig_wc)
                plt.close()
                
            except Exception as e:
                # Fallback: show as bar chart
                st.info(f"📊 Showing word frequency as bar chart")
                
                df_words = pd.DataFrame([
                    {'word': k, 'count': v} for k, v in list(word_freq.items())[:30]
                ])
                df_words = df_words.sort_values('count', ascending=True)
        
                fig_words = px.bar(
                    df_words, y='word', x='count',
                    orientation='h',
                    color='count',
                    color_continuous_scale=['#FFC107', '#4CAF50']
                    )
                fig_words.update_layout(height=600, showlegend=False)
                st.plotly_chart(fig_words, use_container_width=True)
    
            # Show top words table
            with st.expander("📋 View Top 50 Words"):
                df_top_words = pd.DataFrame([
            {'کلمه': k, 'تعداد': v} for k, v in list(word_freq.items())[:50]
        ])
            st.dataframe(df_top_words, use_container_width=True, hide_index=True)
        else:
            st.warning("No text data available for word cloud")

        # ==========================================
        # N-GRAM ANALYSIS
        # ==========================================
        st.markdown("#### 🔗 N-gram Analysis - Common Phrases")
        
        col_ng1, col_ng2 = st.columns(2)
        
        with col_ng1:
            st.markdown("**Bigrams (2-word phrases)**")
            bigrams = analyzer.get_ngram_analysis(n=2, min_freq=3, top_n=20)
            
            if not bigrams.empty:
                fig_bi = px.bar(
                    bigrams.sort_values('count', ascending=True),
                    y='phrase', x='count',
                    orientation='h',
                    color='count',
                    color_continuous_scale=['#2196F3', '#4CAF50']
                )
                fig_bi.update_layout(height=450, showlegend=False, yaxis=dict(tickfont=dict(size=11)))
                st.plotly_chart(fig_bi, use_container_width=True)
        
        with col_ng2:
            st.markdown("**Trigrams (3-word phrases)**")
            trigrams = analyzer.get_ngram_analysis(n=3, min_freq=2, top_n=20)
            
            if not trigrams.empty:
                fig_tri = px.bar(
                    trigrams.sort_values('count', ascending=True),
                    y='phrase', x='count',
                    orientation='h',
                    color='count',
                    color_continuous_scale=['#FF9800', '#F44336']
                )
                fig_tri.update_layout(height=450, showlegend=False, yaxis=dict(tickfont=dict(size=11)))
                st.plotly_chart(fig_tri, use_container_width=True)
        
        # ==========================================
        # KEYWORDS BY RATING
        # ==========================================
        st.markdown("---")
        st.markdown("#### 🎯 Distinctive Keywords by Rating")
        st.caption("Words that appear more frequently in each rating group")
        
        keywords_by_rating = analyzer.get_keywords_by_rating(top_n=15)
        
        if keywords_by_rating:
            col_kw1, col_kw2, col_kw3 = st.columns(3)
            
            with col_kw1:
                st.markdown("""
                <div style="background-color:#D32F2F; padding:10px; border-radius:8px; text-align:center; margin-bottom:10px;">
                    <span style="color:white; font-size:18px;">⭐ 1-2 Stars</span><br>
                    <small style="color:white;">What unhappy customers say</small>
                </div>
                """, unsafe_allow_html=True)
                
                if 'low' in keywords_by_rating and keywords_by_rating['low']:
                    df_low = pd.DataFrame(keywords_by_rating['low'])
                    for _, row in df_low.head(10).iterrows():
                        st.markdown(f"**{row['word']}** ({row['count']})")
                else:
                    st.info("No data")
            
            with col_kw2:
                st.markdown("""
                <div style="background-color:#FF9800; padding:10px; border-radius:8px; text-align:center; margin-bottom:10px;">
                    <span style="color:white; font-size:18px;">⭐⭐⭐ 3 Stars</span><br>
                    <small style="color:white;">What neutral customers say</small>
                </div>
                """, unsafe_allow_html=True)
                
                if 'mid' in keywords_by_rating and keywords_by_rating['mid']:
                    df_mid = pd.DataFrame(keywords_by_rating['mid'])
                    for _, row in df_mid.head(10).iterrows():
                        st.markdown(f"**{row['word']}** ({row['count']})")
                else:
                    st.info("No data")
            
            with col_kw3:
                st.markdown("""
                <div style="background-color:#4CAF50; padding:10px; border-radius:8px; text-align:center; margin-bottom:10px;">
                    <span style="color:white; font-size:18px;">⭐⭐⭐⭐⭐ 4-5 Stars</span><br>
                    <small style="color:white;">What happy customers say</small>
                </div>
                """, unsafe_allow_html=True)
                
                if 'high' in keywords_by_rating and keywords_by_rating['high']:
                    df_high = pd.DataFrame(keywords_by_rating['high'])
                    for _, row in df_high.head(10).iterrows():
                        st.markdown(f"**{row['word']}** ({row['count']})")
                else:
                    st.info("No data")
        
        st.markdown("---")
        
        # ==========================================
        # TOPIC DISCOVERY
        # ==========================================
        st.markdown("#### 🏷️ Topic Discovery")
        st.caption("Main themes found in customer comments")
        
        topics = analyzer.get_topic_keywords(n_topics=5, n_words=10)
        
        if topics:
            cols_topic = st.columns(len(topics))
            
            topic_colors = ['#4CAF50', '#2196F3', '#FF9800', '#9C27B0', '#F44336']
            
            for i, topic in enumerate(topics):
                with cols_topic[i]:
                    color = topic_colors[i % len(topic_colors)]
                    st.markdown(f"""
                    <div style="background-color:{color}; padding:10px; border-radius:8px; text-align:center; min-height:200px;">
                        <b style="color:white; font-size:14px;">{topic['topic']}</b><br>
                        <small style="color:white;">({topic['count']} mentions)</small><br>
                        <hr style="border-color:rgba(255,255,255,0.3);">
                        <small style="color:white;">
                            {'<br>'.join(topic['keywords'][:7])}
                        </small>
                    </div>
                    """, unsafe_allow_html=True)
        else:
            st.info("Not enough data for topic discovery")
        
        # ==========================================
        # SENTIMENT ANALYSIS
        # ==========================================
        st.markdown("---")
        st.markdown("#### 😊😐😠 Sentiment Analysis")
        
        col_sent1, col_sent2 = st.columns(2)
        
        with col_sent1:
            st.markdown("**Sentiment Distribution**")
            
            sentiment_dist = analyzer.get_comment_sentiment_distribution()
            
            if not sentiment_dist.empty:
                sentiment_colors = {
                    'positive': '#4CAF50',
                    'negative': '#D32F2F',
                    'neutral': '#9E9E9E',
                    'mixed': '#FF9800'
                }
                
                fig_sent = px.pie(
                    sentiment_dist,
                    values='count',
                    names='sentiment',
                    color='sentiment',
                    color_discrete_map=sentiment_colors,
                    hole=0.4
                )
                fig_sent.update_traces(textposition='outside', textinfo='label+percent')
                fig_sent.update_layout(height=350, showlegend=False)
                st.plotly_chart(fig_sent, use_container_width=True)
        
        with col_sent2:
            st.markdown("**Rating vs Sentiment Matrix**")
            
            rating_sentiment = analyzer.get_rating_sentiment_matrix()
            
            if not rating_sentiment.empty:
                # Create heatmap
                fig_rs = px.imshow(
                    rating_sentiment.iloc[:-1, :-1],  # Exclude 'All' row/col
                    color_continuous_scale=['#FFEBEE', '#4CAF50'],
                    aspect='auto',
                    text_auto=True
                )
                fig_rs.update_layout(height=350)
                st.plotly_chart(fig_rs, use_container_width=True)
        
        # ==========================================
        # INSIGHTS SUMMARY
        # ==========================================
        st.markdown("---")
        st.markdown("#### 💡 Key Text Mining Insights")
        
        col_ins1, col_ins2, col_ins3 = st.columns(3)
        
        with col_ins1:
            if word_freq:
                top_word = list(word_freq.keys())[0]
                st.metric("🔤 Most Common Word", top_word, f"{word_freq[top_word]} mentions")
        
        with col_ins2:
            if len(bigrams) > 0:
                top_phrase = bigrams.iloc[0]['phrase']
                st.metric("🔗 Most Common Phrase", top_phrase, f"{bigrams.iloc[0]['count']} mentions")
        
        with col_ins3:
            if topics:
                top_topic = topics[0]['topic']
                st.metric("🏷️ Main Topic", top_topic, f"{topics[0]['count']} mentions")
        
    else:
        st.warning("⚠️ No text/comment column found in your data. Text mining requires customer comments.")
        st.info("Expected column names: 'لطفا نظر و انتفادات خود را برای ما بنویسید', 'نظر', 'توضیحات', 'comment'")
# CRITICAL: Increment the index for the next tab
t += 1

## TAB: MACHINE LEARNING
with tabs[t]:  # Adjust index as needed
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("### 🤖 Machine Learning Analysis")
    st.caption("Predictive models and advanced pattern discovery")
    
    # Initialize ML Analyzer
    from config import COLS
    ml_analyzer = ShilaMLAnalyzer(st.session_state.df, COLS)
    ml_summary = ml_analyzer.get_ml_summary()
    
    if not ml_summary['ml_available']:
        st.error("⚠️ scikit-learn not installed. Run: `pip install scikit-learn`")
        st.stop()
    
    # ==========================================
    # DYNAMIC ML SUB-TABS
    # ==========================================
    # We remove Detractor and Churn prediction for SnappFood 
    # because they rely on NPS logic.
    ml_tab_titles = ["👥 Customer Clustering", "🔗 Association Rules", "🚨 Anomaly Detection"]
    if not st.session_state.get('is_snappfood', False):
        ml_tab_titles.insert(0, "🎯 Detractor Prediction")
        ml_tab_titles.append("📉 Churn Prediction")
    
    ml_tabs = st.tabs(ml_tab_titles)
    
    # Local counter for ML sub-tabs
    m = 0
    
    # 1. SUB-TAB: DETRACTOR PREDICTION (Conditional)
    if not st.session_state.get('is_snappfood', False):
        with ml_tabs[m]:
            st.markdown("#### 🎯 Predict Potential Detractors")
            st.caption("Identify customers likely to give low NPS scores before they do")
        
            if st.button("🚀 Train Detractor Model", key="train_detractor"):
                with st.spinner("Training model..."):
                    results = ml_analyzer.train_detractor_model()
            
                if 'error' in results:
                    st.error(results['error'])
                else:
                    # Model Performance
                    st.markdown("##### 📊 Model Performance")
                
                    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                    with col_m1:
                        st.metric("Accuracy", f"{results['accuracy']*100:.1f}%")
                    with col_m2:
                        st.metric("Precision", f"{results['precision']*100:.1f}%")
                    with col_m3:
                        st.metric("Recall", f"{results['recall']*100:.1f}%")
                    with col_m4:
                        st.metric("F1 Score", f"{results['f1_score']*100:.1f}%")
                
                    st.markdown("---")
                
                    col_conf, col_feat = st.columns(2)
                
                    with col_conf:
                        st.markdown("##### 🔢 Confusion Matrix")
                        cm = results['confusion_matrix']
                    
                        fig_cm = go.Figure(data=go.Heatmap(
                            z=cm,
                            x=['Predicted: Not Detractor', 'Predicted: Detractor'],
                            y=['Actual: Not Detractor', 'Actual: Detractor'],
                            text=cm,
                            texttemplate='%{text}',
                            colorscale=['#E8F5E9', '#4CAF50'],
                            showscale=False
                        ))
                        fig_cm.update_layout(height=300, margin=dict(l=10, r=10, t=30, b=10))
                        st.plotly_chart(fig_cm, use_container_width=True)
                        
                    with col_feat:
                        st.markdown("##### 📈 Feature Importance")
                            
                        importance_df = pd.DataFrame(results['feature_importance'])
                            
                        fig_imp = px.bar(
                            importance_df.head(10).sort_values('importance'),
                            y='feature', x='importance',
                            orientation='h',
                            color='importance',
                            color_continuous_scale=['#FFC107', '#4CAF50']
                        )
                        fig_imp.update_layout(height=300, showlegend=False, margin=dict(l=10, r=10, t=30, b=10))
                        st.plotly_chart(fig_imp, use_container_width=True)
                
                        st.success(f"✅ Model trained! Detractor rate: {results['detractor_rate']}%")
        
                st.markdown("---")
        
                # High Risk Customers
                st.markdown("##### 🚨 High Risk Customers")
        
                if st.button("🔍 Find High Risk Customers", key="find_risk"):
                    with st.spinner("Analyzing..."):
                        high_risk = ml_analyzer.predict_detractor_risk(top_n=50)
            
                    if len(high_risk) > 0:
                        # Risk distribution
                        risk_counts = high_risk['risk_level'].value_counts()
                        
                        col_r1, col_r2, col_r3 = st.columns(3)
                        with col_r1:
                            high_count = risk_counts.get('High', 0)
                            st.metric("🔴 High Risk", high_count)
                        with col_r2:
                            med_count = risk_counts.get('Medium', 0)
                            st.metric("🟠 Medium Risk", med_count)
                        with col_r3:
                            low_count = risk_counts.get('Low', 0)
                            st.metric("🟢 Low Risk", low_count)
                
                        st.dataframe(high_risk, use_container_width=True, hide_index=True)
                    else:
                        st.info("No risk data available")
            m += 1
    
    # 2. SUB-TAB: CLUSTERING (Universal)
    with ml_tabs[m]:
        st.markdown("#### 👥 Customer Clustering")
        st.caption("Discover natural customer segments using K-Means clustering")
        
        n_clusters = st.slider("Number of Clusters", 2, 8, 5)
        
        if st.button("🔬 Perform Clustering", key="run_cluster"):
            with st.spinner("Clustering customers..."):
                cluster_results = ml_analyzer.perform_clustering(n_clusters=n_clusters)
            
            if 'error' in cluster_results:
                st.error(cluster_results['error'])
            else:
                # Cluster Stats
                st.markdown("##### 📊 Cluster Profiles")
                
                cluster_df = pd.DataFrame(cluster_results['cluster_stats'])
                
                # Display as cards
                cols = st.columns(len(cluster_df))
                
                cluster_colors = ['#4CAF50', '#8BC34A', '#FFC107', '#FF9800', '#F44336', '#9C27B0', '#2196F3', '#00BCD4']
                
                for i, (_, row) in enumerate(cluster_df.iterrows()):
                    with cols[i]:
                        color = cluster_colors[i % len(cluster_colors)]
                        st.markdown(f"""
                        <div style="background-color:{color}; padding:15px; border-radius:10px; text-align:center; color:white;">
                            <b style="font-size:14px;">{row.get('cluster_name', f'Cluster {i}')}</b><br>
                            <span style="font-size:24px;"><b>{row['size']:,}</b></span><br>
                            <small>({row['percentage']}%)</small><br>
                            <hr style="border-color:rgba(255,255,255,0.3);">
                            <small>Rating: {row.get('avg_rating', 'N/A')}</small><br>
                            <small>NPS: {row.get('avg_nps', 'N/A')}</small>
                        </div>
                        """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                col_elbow, col_pca = st.columns(2)
                
                with col_elbow:
                    st.markdown("##### 📉 Elbow Method")
                    elbow = cluster_results['elbow_data']
                    
                    fig_elbow = px.line(
                        x=elbow['k'], y=elbow['inertia'],
                        markers=True,
                        labels={'x': 'Number of Clusters (K)', 'y': 'Inertia'}
                    )
                    fig_elbow.update_layout(height=300)
                    st.plotly_chart(fig_elbow, use_container_width=True)
                
                with col_pca:
                    st.markdown("##### 🎯 Cluster Visualization (PCA)")
                    pca = cluster_results['pca_data']
                    
                    fig_pca = px.scatter(
                        x=pca['x'], y=pca['y'],
                        color=[str(c) for c in pca['cluster']],
                        labels={'x': 'PC1', 'y': 'PC2', 'color': 'Cluster'},
                        color_discrete_sequence=cluster_colors
                    )
                    fig_pca.update_layout(height=300)
                    st.plotly_chart(fig_pca, use_container_width=True)
                
                # Full stats table
                with st.expander("📋 View Full Cluster Statistics"):
                    st.dataframe(cluster_df, use_container_width=True, hide_index=True)
    m += 1
    
    # 3. SUB-TAB: Association rules
    with ml_tabs[m]:
        st.markdown("#### 🔗 Association Rules")
        st.caption("Discover which issues frequently occur together")
        
        if not ml_summary.get('mlxtend_available', False):
            st.warning("⚠️ mlxtend not installed. Run: `pip install mlxtend`")
        else:
            col_params1, col_params2 = st.columns(2)
            with col_params1:
                min_support = st.slider("Minimum Support", 0.005, 0.1, 0.01, 0.005)
            with col_params2:
                min_confidence = st.slider("Minimum Confidence", 0.1, 0.8, 0.3, 0.05)
            
            if st.button("🔍 Find Association Rules", key="find_rules"):
                with st.spinner("Mining rules..."):
                    rules_results = ml_analyzer.get_association_rules(
                        min_support=min_support,
                        min_confidence=min_confidence
                    )
                
                if 'error' in rules_results:
                    st.error(rules_results['error'])
                elif not rules_results.get('rules'):
                    st.info("No association rules found with the current thresholds. Try lowering Support or Confidence.")
                else:
                    st.success(f"Found {len(rules_results['rules'])} rules from {rules_results['total_transactions']:,} transactions")
                    
                    st.markdown("##### 📜 Top Association Rules")
                    st.caption("'If X, then Y' - What issues appear together")
                    
                    for rule in rules_results['rules'][:10]:
                        lift_color = '#4CAF50' if rule['lift'] > 1.5 else ('#FFC107' if rule['lift'] > 1 else '#9E9E9E')
                        st.markdown(f"""
                        <div style="background-color:#f5f5f5; padding:10px; border-radius:8px; margin:5px 0; border-left:4px solid {lift_color};">
                            <b>IF</b> {rule['if']}<br>
                            <b>THEN</b> {rule['then']}<br>
                            <small style="color:#666;">
                                Support: {rule['support']:.1%} | 
                                Confidence: {rule['confidence']:.1%} | 
                                Lift: <span style="color:{lift_color}; font-weight:bold;">{rule['lift']:.2f}</span>
                            </small>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    st.markdown("---")
                    
                    st.markdown("##### 📊 Frequent Issue Combinations")
                    itemsets_df = pd.DataFrame(rules_results['frequent_itemsets'])
                    if not itemsets_df.empty:
                        st.dataframe(itemsets_df, use_container_width=True, hide_index=True)

    # Increment the ML sub-tab index
    m += 1
    
    # 4. SUB-TAB: ANOMALY DETECTION
    with ml_tabs[m]:
        st.markdown("#### 🚨 Anomaly Detection")
        st.caption("Find unusual patterns that may indicate fraud, errors, or system issues")
        
        contamination = st.slider("Expected Anomaly Rate", 0.01, 0.15, 0.05, 0.01)
        
        if st.button("🔍 Detect Anomalies", key="detect_anomaly"):
            with st.spinner("Analyzing patterns..."):
                anomaly_results = ml_analyzer.detect_anomalies(contamination=contamination)
            
            if 'error' in anomaly_results:
                st.error(anomaly_results['error'])
            else:
                stats = anomaly_results['stats']
                
                # Summary metrics
                col_a1, col_a2, col_a3, col_a4 = st.columns(4)
                
                with col_a1:
                    st.metric("🚨 Anomalies Found", stats['total_anomalies'])
                with col_a2:
                    st.metric("📊 Anomaly Rate", f"{stats['anomaly_rate']}%")
                with col_a3:
                    if 'anomaly_avg_rating' in stats:
                        st.metric("⭐ Anomaly Avg Rating", stats['anomaly_avg_rating'])
                with col_a4:
                    if 'normal_avg_rating' in stats:
                        st.metric("⭐ Normal Avg Rating", stats['normal_avg_rating'])
                
                st.markdown("---")
                
                # Anomaly Types
                if anomaly_results['anomaly_types']:
                    st.markdown("##### 🏷️ Anomaly Types Detected")
                    
                    cols_type = st.columns(len(anomaly_results['anomaly_types']))
                    
                    for i, atype in enumerate(anomaly_results['anomaly_types']):
                        with cols_type[i]:
                            st.markdown(f"""
                            <div style="background-color:#FFF3E0; padding:15px; border-radius:10px; text-align:center; border:2px solid #FF9800;">
                                <span style="font-size:28px;">{atype['icon']}</span><br>
                                <b>{atype['type']}</b><br>
                                <span style="font-size:24px; color:#FF9800;"><b>{atype['count']}</b></span><br>
                                <small>{atype['description']}</small>
                            </div>
                            """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Top anomalies table
                st.markdown("##### 📋 Top Anomalies to Review")
                anomalies_df = pd.DataFrame(anomaly_results['top_anomalies'])
                
                if not anomalies_df.empty:
                    st.dataframe(anomalies_df, use_container_width=True, hide_index=True)
                else:
                    st.info("No specific anomaly records to display.")

    # Increment the local ML counter
    m += 1
    

    # 5. SUB-TAB: CHURN PREDICTION
    if not st.session_state.get('is_snappfood', False):
        with ml_tabs[m]:
            st.markdown("#### 📉 Churn Prediction")
            st.caption("Predict which customers are likely to stop ordering")
        
            st.info("💡 **Note:** True churn prediction requires repeat customer data (customer ID + order history). This model uses a proxy based on rating, NPS, and issues.")
        
            if st.button("🚀 Train Churn Model", key="train_churn"):
                with st.spinner("Training model..."):
                    churn_results = ml_analyzer.train_churn_model()
            
                if 'error' in churn_results:
                    st.error(churn_results['error'])
                else:
                    # Model Performance
                    st.markdown("##### 📊 Model Performance")
                
                    col_c1, col_c2, col_c3, col_c4 = st.columns(4)
                    with col_c1:
                        st.metric("Accuracy", f"{churn_results['accuracy']*100:.1f}%")
                    with col_c2:
                        st.metric("Precision", f"{churn_results['precision']*100:.1f}%")
                    with col_c3:
                        st.metric("Recall", f"{churn_results['recall']*100:.1f}%")
                    with col_c4:
                        st.metric("Churn Rate", f"{churn_results['churn_rate']}%")
                
                    st.markdown("---")
                
                    col_cm, col_fi = st.columns(2)
                
                    with col_cm:
                        st.markdown("##### 🔢 Confusion Matrix")
                        cm = churn_results['confusion_matrix']
                        
                        fig_cm = go.Figure(data=go.Heatmap(
                            z=cm,
                            x=['Stay', 'Churn'],
                            y=['Actual: Stay', 'Actual: Churn'],
                            text=cm,
                            texttemplate='%{text}',
                            colorscale=['#E3F2FD', '#2196F3'],
                            showscale=False
                        ))
                        fig_cm.update_layout(height=300, margin=dict(l=10, r=10, t=30, b=10))
                        st.plotly_chart(fig_cm, use_container_width=True)
                        
                        with col_fi:
                            st.markdown("##### 📈 Feature Importance")
                            
                            importance_df = pd.DataFrame(churn_results['feature_importance'])
                            
                            fig_imp = px.bar(
                                importance_df.sort_values('importance'),
                                y='feature', x='importance',
                                orientation='h',
                                color='importance',
                                color_continuous_scale=['#BBDEFB', '#2196F3']
                            )
                            fig_imp.update_layout(height=300, showlegend=False, margin=dict(l=10, r=10, t=30, b=10))
                            st.plotly_chart(fig_imp, use_container_width=True)
                            
                st.markdown("---")
        
                # High Churn Risk
                st.markdown("##### 📉 High Churn Risk Customers")
        
                if st.button("🔍 Find Churn Risk Customers", key="find_churn"):
                    with st.spinner("Analyzing..."):
                        churn_risk = ml_analyzer.predict_churn_risk(top_n=50)
            
                if len(churn_risk) > 0:
                    risk_counts = churn_risk['churn_level'].value_counts()
                
                    col_cr1, col_cr2, col_cr3 = st.columns(3)
                    with col_cr1:
                        st.metric("🔴 High Churn Risk", risk_counts.get('High', 0))
                    with col_cr2:
                        st.metric("🟠 Medium Risk", risk_counts.get('Medium', 0))
                    with col_cr3:
                        st.metric("🟢 Low Risk", risk_counts.get('Low', 0))
                
                        st.dataframe(churn_risk, use_container_width=True, hide_index=True)
                else:
                    st.info("No churn risk data available")

        # Increment local counter
        m += 1

    # FINAL: Increment the main tab counter so your app stays perfectly synced
    t += 1

# --- EXPORT FOOTER ---
st.markdown("---")
st.markdown(f"### {L('export_section')}")
c_exp_1, c_exp_2, c_exp_3 = st.columns(3)

with c_exp_1:
    if st.button(f"📥 {L('export_excel')}", use_container_width=True):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fp = os.path.join(REPORTS_DIR, f"full_analysis_{ts}.xlsx")
        
        with st.spinner("Generating comprehensive Excel report with charts..."):
            try:
                # Create workbook
                from openpyxl import Workbook
                from openpyxl.drawing.image import Image as XLImage
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.chart import BarChart, LineChart, PieChart, Reference
                import io
                
                wb = Workbook()
                
                # Helper function to add dataframe to sheet
                def df_to_sheet(ws, df, start_row=1):
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            if r_idx == start_row:  # Header row
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")
                    return r_idx + 2  # Return next available row
                
                # Helper to save plotly chart as image bytes
                def fig_to_image_bytes(fig, width=700, height=400):
                    img_bytes = fig.to_image(format="png", width=width, height=height, scale=2)
                    return io.BytesIO(img_bytes)
                
                # ==========================================
                # SHEET 1: DASHBOARD SUMMARY WITH CHARTS
                # ==========================================
                ws = wb.active
                ws.title = "📊 Dashboard"
                
                # Title
                ws['A1'] = "🍕 Shila Restaurant - QFD Analysis Report"
                ws['A1'].font = Font(bold=True, size=18, color="1A1F36")
                ws.merge_cells('A1:F1')
                
                ws['A2'] = f"Generated: {ts}"
                ws['A2'].font = Font(italic=True, color="697386")
                
                # KPI Summary
                ws['A4'] = "📈 Key Performance Indicators"
                ws['A4'].font = Font(bold=True, size=14)
                
                kpi_data = [
                    ['Metric', 'Value', 'Status'],
                    ['NPS Score', kpis['nps_score'], '🟢 Good' if kpis['nps_score'] > 30 else ('🟡 OK' if kpis['nps_score'] > 0 else '🔴 Bad')],
                    ['Average Rating', f"{kpis['avg_rating']} / 5", '🟢 Good' if kpis['avg_rating'] >= 4 else ('🟡 OK' if kpis['avg_rating'] >= 3 else '🔴 Bad')],
                    ['Total Orders', kpis['total_orders'], '-'],
                    ['Promoters', kpis['promoters'], '😊'],
                    ['Passives', kpis['passives'], '😐'],
                    ['Detractors', kpis['detractors'], '😠'],
                    ['Response Rate', f"{kpis['response_rate']}%", '🟢 Good' if kpis['response_rate'] > 50 else '🟡 Low'],
                ]
                for r_idx, row in enumerate(kpi_data, 5):
                    for c_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        if r_idx == 5:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                
                # Add Rating Distribution Chart
                rd = analyzer.get_rating_distribution()
                if len(rd) > 0:
                    fig_rating = px.bar(
                        rd, x='Rating', y='Count', 
                        color='Rating',
                        color_continuous_scale=['#D32F2F','#FF9800','#FFEB3B','#8BC34A','#4CAF50'],
                        title='Rating Distribution'
                    )
                    fig_rating.update_layout(
                        paper_bgcolor='white', plot_bgcolor='white',
                        font=dict(family="Arial", size=12),
                        showlegend=False
                    )
                    img_bytes = fig_to_image_bytes(fig_rating, 500, 350)
                    img = XLImage(img_bytes)
                    ws.add_image(img, 'E5')
                
                # ==========================================
                # SHEET 2: NPS ANALYSIS WITH CHART
                # ==========================================
                ws2 = wb.create_sheet("📈 NPS Analysis")
                
                ws2['A1'] = "NPS Score Analysis"
                ws2['A1'].font = Font(bold=True, size=16)
                
                nd = analyzer.get_nps_distribution()
                if len(nd) > 0:
                    next_row = df_to_sheet(ws2, nd, 3)
                    
                    # NPS Distribution Chart
                    fig_nps = px.bar(
                        nd, x='NPS', y='Count', color='Segment',
                        color_discrete_map={'Promoter':'#4CAF50','Passive':'#FF9800','Detractor':'#D32F2F'},
                        title='NPS Score Distribution'
                    )
                    fig_nps.update_layout(paper_bgcolor='white', plot_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_nps, 600, 400)
                    img = XLImage(img_bytes)
                    ws2.add_image(img, 'F3')
                    
                    # NPS Pie Chart
                    segment_counts = nd.groupby('Segment')['Count'].sum().reset_index()
                    fig_pie = px.pie(
                        segment_counts, values='Count', names='Segment',
                        color='Segment',
                        color_discrete_map={'Promoter':'#4CAF50','Passive':'#FF9800','Detractor':'#D32F2F'},
                        title='NPS Segments'
                    )
                    fig_pie.update_layout(paper_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_pie, 450, 400)
                    img = XLImage(img_bytes)
                    ws2.add_image(img, 'F20')
                
                # ==========================================
                # SHEET 3: PARETO ANALYSIS WITH CHART
                # ==========================================
                ws3 = wb.create_sheet("📊 Pareto Analysis")
                
                ws3['A1'] = "Pareto Analysis - Issues by Impact"
                ws3['A1'].font = Font(bold=True, size=16)
                
                pareto = analyzer.get_pareto_analysis()
                if len(pareto) > 0:
                    df_to_sheet(ws3, pareto, 3)
                    
                    # Pareto Chart
                    fig_pareto = make_subplots(specs=[[{"secondary_y": True}]])
                    fig_pareto.add_trace(go.Bar(
                        x=pareto['tag'].head(15), y=pareto['total_damage'].head(15),
                        name='Impact Score', marker_color='#D32F2F', opacity=0.85
                    ), secondary_y=False)
                    fig_pareto.add_trace(go.Scatter(
                        x=pareto['tag'].head(15), y=pareto['cumulative_pct'].head(15),
                        name='Cumulative %', mode='lines+markers',
                        line=dict(color='#1A1F36', width=2)
                    ), secondary_y=True)
                    fig_pareto.add_hline(y=80, line_dash="dash", line_color="#4CAF50", secondary_y=True)
                    fig_pareto.update_layout(
                        title='Pareto Chart - Top Issues by Rating Damage',
                        paper_bgcolor='white', plot_bgcolor='white',
                        xaxis_tickangle=-45
                    )
                    img_bytes = fig_to_image_bytes(fig_pareto, 800, 450)
                    img = XLImage(img_bytes)
                    ws3.add_image(img, 'H3')
                
                # ==========================================
                # SHEET 4: KANO MODEL WITH CHART
                # ==========================================
                ws4 = wb.create_sheet("🎨 Kano Model")
                
                ws4['A1'] = "Kano Model Classification"
                ws4['A1'].font = Font(bold=True, size=16)
                
                kano = analyzer.get_kano_analysis()
                if len(kano) > 0:
                    df_to_sheet(ws4, kano, 3)
                    
                    # Kano Scatter Chart
                    fig_kano = px.scatter(
                        kano, x='lift_as_strength', y='drop_as_weakness',
                        color='kano_type', hover_name='attribute',
                        size='strength_mentions', size_max=40,
                        color_discrete_map={'Must-Be':'#D32F2F', 'Performance':'#FFB020', 'Delighter':'#4CAF50'},
                        title='Kano Model - Feature Classification'
                    )
                    fig_kano.add_hline(y=0.5, line_dash="dot", line_color="#E1E4E8")
                    fig_kano.add_vline(x=0.3, line_dash="dot", line_color="#E1E4E8")
                    fig_kano.update_layout(paper_bgcolor='white', plot_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_kano, 700, 500)
                    img = XLImage(img_bytes)
                    ws4.add_image(img, 'H3')
                
                # ==========================================
                # SHEET 5: BRANCH ANALYSIS WITH CHART
                # ==========================================
                ws5 = wb.create_sheet("🏪 Branch Analysis")
                
                ws5['A1'] = "Branch Performance Comparison"
                ws5['A1'].font = Font(bold=True, size=16)
                
                br_stats, br_issues = analyzer.get_branch_analysis()
                if len(br_stats) > 0:
                    df_to_sheet(ws5, br_stats.round(2), 3)
                    
                    # Branch Comparison Chart
                    br_sorted = br_stats.sort_values('avg_rating', ascending=True)
                    fig_branch = px.bar(
                        br_sorted, x='branch', y='rating_vs_avg',
                        color='rating_vs_avg',
                        color_continuous_scale=['#D32F2F', '#FFB020', '#4CAF50'],
                        color_continuous_midpoint=0,
                        title='Branch Performance vs Average'
                    )
                    fig_branch.update_layout(paper_bgcolor='white', plot_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_branch, 700, 400)
                    img = XLImage(img_bytes)
                    ws5.add_image(img, 'I3')
                
                # ==========================================
                # SHEET 6: PRODUCT ANALYSIS
                # ==========================================
                ws6 = wb.create_sheet("🍔 Products")
                
                ws6['A1'] = "Product Performance"
                ws6['A1'].font = Font(bold=True, size=16)
                
                products = analyzer.get_product_analysis()
                if len(products) > 0:
                    df_to_sheet(ws6, products, 3)
                    
                    # Product Chart - HORIZONTAL for Persian text
                    top_products = products.head(15).sort_values('avg_rating', ascending=True)
    
                    fig_prod = px.bar(
                        top_products,
                        y='product',
                        x='avg_rating',
                        orientation='h',
                        color='avg_rating',
                        color_continuous_scale=['#D32F2F', '#FF9800', '#FFEB3B', '#8BC34A', '#4CAF50'],
                        text='avg_rating',
                        title='Top Products by Rating'
                    )
                    
                    fig_prod.update_traces(texttemplate='%{text:.2f}', textposition='outside')
                    fig_prod.update_layout(
                        paper_bgcolor='white', 
                        plot_bgcolor='white',
                        height=500,
                        yaxis=dict(automargin=True),
                        xaxis=dict(range=[0, 5.5]),
                        showlegend=False
                    )
                    img_bytes = fig_to_image_bytes(fig_prod, 700, 500)
                    img = XLImage(img_bytes)
                    ws6.add_image(img, 'G3')
                
                # ==========================================
                # SHEET 7: CUSTOMER SEGMENTS WITH CHART
                # ==========================================
                ws7 = wb.create_sheet("🎯 Customer Segments")
                
                ws7['A1'] = "Customer Segmentation Analysis"
                ws7['A1'].font = Font(bold=True, size=16)
                
                recovery = analyzer.get_recovery_opportunities()
                if len(recovery) > 0:
                    df_to_sheet(ws7, recovery, 3)
                    
                    # Segment Pie Chart
                    fig_seg = px.pie(
                        recovery, values='count', names='segment',
                        color='segment',
                        color_discrete_map={
                            'Happy':'#4CAF50', 'Neutral':'#9E9E9E',
                            'Recovery':'#FF9800', 'Silent Churner':'#FF5722', 'At Risk':'#D32F2F'
                        },
                        title='Customer Segments Distribution'
                    )
                    fig_seg.update_layout(paper_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_seg, 500, 400)
                    img = XLImage(img_bytes)
                    ws7.add_image(img, 'G3')
                
                # ==========================================
                # SHEET 8: ISSUE CATEGORIES
                # ==========================================
                ws8 = wb.create_sheet("⚠️ Issue Categories")
                
                ws8['A1'] = "Issue Category Impact Analysis"
                ws8['A1'].font = Font(bold=True, size=16)
                
                issue_cats = analyzer.get_issue_category_analysis()
                if len(issue_cats) > 0:
                    df_to_sheet(ws8, issue_cats, 3)
                    
                    # Issue Category Chart
                    fig_cats = px.bar(
                        issue_cats, x='category_fa', y='rating_impact',
                        color='rating_impact',
                        color_continuous_scale=['#4CAF50', '#FFB020', '#D32F2F'],
                        title='Rating Impact by Issue Category'
                    )
                    fig_cats.update_layout(paper_bgcolor='white', plot_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_cats, 500, 350)
                    img = XLImage(img_bytes)
                    ws8.add_image(img, 'J3')
                
                # ==========================================
                # SHEET 9: ASPECT SENTIMENT WITH CHART
                # ==========================================
                ws9 = wb.create_sheet("🎭 Aspect Sentiment")
                
                ws9['A1'] = "Aspect-Based Sentiment Analysis"
                ws9['A1'].font = Font(bold=True, size=16)
                
                aspects = analyzer.get_aspect_sentiment()
                if len(aspects) > 0:
                    df_to_sheet(ws9, aspects, 3)
                    
                    # Sentiment Chart
                    fig_aspect = px.bar(
                        aspects, y='aspect', x='sentiment_score', orientation='h',
                        color='sentiment_score',
                        color_continuous_scale=['#D32F2F', '#FFEB3B', '#4CAF50'],
                        color_continuous_midpoint=0,
                        title='Aspect Sentiment Scores'
                    )
                    fig_aspect.update_layout(paper_bgcolor='white', plot_bgcolor='white')
                    img_bytes = fig_to_image_bytes(fig_aspect, 600, 400)
                    img = XLImage(img_bytes)
                    ws9.add_image(img, 'H3')
                
                # ==========================================
                # SHEET 10: DAILY TRENDS WITH CHART
                # ==========================================
                ws10 = wb.create_sheet("📅 Daily Trends")
                
                ws10['A1'] = "Daily Performance Trends"
                ws10['A1'].font = Font(bold=True, size=16)
                
                daily = analyzer.get_daily_trends()
                if len(daily) > 0:
                    df_to_sheet(ws10, daily.round(2), 3)
                    
                    # Trend Chart
                    fig_trend = make_subplots(rows=2, cols=1, subplot_titles=('Rating Trend', 'Order Volume'))
                    fig_trend.add_trace(go.Scatter(
                        x=daily['date'], y=daily['avg_rating'],
                        mode='lines', name='Daily', line=dict(color='#E1E4E8', width=1)
                    ), row=1, col=1)
                    fig_trend.add_trace(go.Scatter(
                        x=daily['date'], y=daily['rating_7day_avg'],
                        mode='lines', name='7-Day Avg', line=dict(color='#2196F3', width=3)
                    ), row=1, col=1)
                    fig_trend.add_trace(go.Bar(
                        x=daily['date'], y=daily['order_count'],
                        name='Orders', marker_color='#4CAF50', opacity=0.6
                    ), row=2, col=1)
                    fig_trend.update_layout(
                        paper_bgcolor='white', plot_bgcolor='white',
                        height=500, showlegend=True
                    )
                    img_bytes = fig_to_image_bytes(fig_trend, 900, 500)
                    img = XLImage(img_bytes)
                    ws10.add_image(img, 'H3')
                
                # ==========================================
                # SHEET 11: MONTHLY TRENDS WITH CHART
                # ==========================================
                ws11 = wb.create_sheet("📅 Monthly Trends")
                
                ws11['A1'] = "Month-over-Month Analysis"
                ws11['A1'].font = Font(bold=True, size=16)
                
                mom = analyzer.get_mom_comparison()
                if len(mom) > 0:
                    df_to_sheet(ws11, mom, 3)
                    
                    # MoM Chart
                    fig_mom = make_subplots(specs=[[{"secondary_y": True}]])
                    fig_mom.add_trace(go.Bar(
                        x=mom['year_month'], y=mom['order_count'],
                        name='Orders', marker_color='#4CAF50', opacity=0.6
                    ), secondary_y=False)
                    fig_mom.add_trace(go.Scatter(
                        x=mom['year_month'], y=mom['avg_rating'],
                        name='Avg Rating', mode='lines+markers',
                        line=dict(color='#2196F3', width=3)
                    ), secondary_y=True)
                    fig_mom.update_layout(
                        title='Monthly Performance',
                        paper_bgcolor='white', plot_bgcolor='white'
                    )
                    img_bytes = fig_to_image_bytes(fig_mom, 700, 400)
                    img = XLImage(img_bytes)
                    ws11.add_image(img, 'H3')
                
                # ==========================================
                # SHEET 12: BRANCH-PRODUCT MATRIX (HEATMAP)
                # ==========================================
                ws12 = wb.create_sheet("🔥 Branch-Product Matrix")
                
                ws12['A1'] = "Branch × Product Performance Matrix"
                ws12['A1'].font = Font(bold=True, size=16)

                matrix = analyzer.get_branch_product_matrix()
                if len(matrix) > 0:
                    # Write matrix manually (with index as first column)
                    # Header row
                    ws12.cell(row=3, column=1, value="Branch").font = Font(bold=True)
                    for c_idx, col_name in enumerate(matrix.columns, 2):
                        cell = ws12.cell(row=3, column=c_idx, value=col_name)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
                    # Data rows
                    for r_idx, (branch, row_data) in enumerate(matrix.iterrows(), 4):
                        ws12.cell(row=r_idx, column=1, value=branch)
                        for c_idx, value in enumerate(row_data, 2):
                            if pd.notna(value):
                                ws12.cell(row=r_idx, column=c_idx, value=round(value, 2))
    
                    # Heatmap chart
                    fig_heat = px.imshow(
                        matrix,
                        color_continuous_scale=['#D32F2F', '#FFEB3B', '#4CAF50'],
                        aspect='auto', text_auto='.2f',
                        title='Branch-Product Rating Heatmap'
                    )
                    fig_heat.update_layout(paper_bgcolor='white', height=500)
                    img_bytes = fig_to_image_bytes(fig_heat, 800, 500)
                    img = XLImage(img_bytes)
                    ws12.add_image(img, f'A{len(matrix)+8}')
                
                # ==========================================
                # SHEET 13: CO-OCCURRENCE
                # ==========================================
                ws13 = wb.create_sheet("🔗 Issue Co-occurrence")
                
                ws13['A1'] = "Issue Co-occurrence Analysis"
                ws13['A1'].font = Font(bold=True, size=16)
                
                cooccur = analyzer.get_cooccurrence(20)
                if len(cooccur) > 0:
                    df_to_sheet(ws13, cooccur, 3)
                
                # ==========================================
                # SHEET 14: TOP ISSUES
                # ==========================================
                ws14 = wb.create_sheet("🚨 Top Issues")
                ws14['A1'] = "Top Issues"
                ws14['A1'].font = Font(bold=True, size=16)
                issues = analyzer.get_top_issues(20)
                if len(issues) > 0:
                    df_to_sheet(ws14, issues, 3)
                
                # ==========================================
                # SHEET 15: TOP STRENGTHS
                # ==========================================
                ws15 = wb.create_sheet("🏆 Top Strengths")
                ws15['A1'] = "Top Strengths"
                ws15['A1'].font = Font(bold=True, size=16)
                strengths = analyzer.get_top_strengths(20)
                if len(strengths) > 0:
                    df_to_sheet(ws15, strengths, 3)
                
                # ==========================================
                # SHEET 16: RAW DATA
                # ==========================================
                ws16 = wb.create_sheet("📁 Raw Data")
                ws16['A1'] = "Original Dataset"
                ws16['A1'].font = Font(bold=True, size=16)
                df_to_sheet(ws16, st.session_state.df, 3)
                
                # ==========================================
                # SHEET 17: WORD FREQUENCY
                # ==========================================
                ws17 = wb.create_sheet("📝 Word Frequency")
                ws17['A1'] = "Word Frequency Analysis"
                ws17['A1'].font = Font(bold=True, size=16)

                word_freq = analyzer.get_word_frequency(min_freq=5, top_n=100)
                if word_freq:
                    # Headers
                    ws17['A3'] = "Word"
                    ws17['B3'] = "Count"
                    ws17['A3'].font = Font(bold=True, color="FFFFFF")
                    ws17['B3'].font = Font(bold=True, color="FFFFFF")
                    ws17['A3'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                    ws17['B3'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    
                # Data
                for i, (word, count) in enumerate(word_freq.items(), 4):
                    ws17.cell(row=i, column=1, value=word)
                    ws17.cell(row=i, column=2, value=count)
    
                    # Word Cloud Chart (Bar chart as alternative)
                    df_wf = pd.DataFrame([{'word': k, 'count': v} for k, v in list(word_freq.items())[:20]])
                    df_wf = df_wf.sort_values('count', ascending=True)
    
                    fig_wf = px.bar(
                        df_wf, y='word', x='count',
                        orientation='h',
                        color='count',
                        color_continuous_scale=['#FFC107', '#4CAF50'],
                        title='Top 20 Words'
                        )
                    fig_wf.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=500, showlegend=False)
                    img_bytes = fig_to_image_bytes(fig_wf, 600, 500)
                    img = XLImage(img_bytes)
                    ws17.add_image(img, 'D3')

                # ==========================================
                # SHEET 18: N-GRAM ANALYSIS
                # ==========================================
                ws18 = wb.create_sheet("🔗 N-gram Analysis")                
                ws18['A1'] = "N-gram Analysis - Common Phrases"
                ws18['A1'].font = Font(bold=True, size=16)

                # Bigrams
                ws18['A3'] = "Bigrams (2-word phrases)"
                ws18['A3'].font = Font(bold=True, size=14)

                bigrams = analyzer.get_ngram_analysis(n=2, min_freq=3, top_n=30)
                if len(bigrams) > 0:
                    ws18['A4'] = "Phrase"
                    ws18['B4'] = "Count"
                    ws18['A4'].font = Font(bold=True, color="FFFFFF")
                    ws18['B4'].font = Font(bold=True, color="FFFFFF")
                    ws18['A4'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                    ws18['B4'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    
                    for i, (_, row) in enumerate(bigrams.iterrows(), 5):
                        ws18.cell(row=i, column=1, value=row['phrase'])
                        ws18.cell(row=i, column=2, value=row['count'])
                    
                    # Bigram Chart
                    bi_chart = bigrams.head(15).sort_values('count', ascending=True)
                    fig_bi = px.bar(
                        bi_chart, y='phrase', x='count',
                        orientation='h',
                        color='count',
                        color_continuous_scale=['#2196F3', '#4CAF50'],
                        title='Top 15 Bigrams'
                    )
                    fig_bi.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=400, showlegend=False)
                    img_bytes = fig_to_image_bytes(fig_bi, 500, 400)
                    img = XLImage(img_bytes)
                    ws18.add_image(img, 'D3')

                # Trigrams
                trigrams = analyzer.get_ngram_analysis(n=3, min_freq=2, top_n=30)
                if len(trigrams) > 0:
                    start_row = max(len(bigrams) + 7, 38)
    
                    ws18.cell(row=start_row, column=1, value="Trigrams (3-word phrases)")
                    ws18.cell(row=start_row, column=1).font = Font(bold=True, size=14)
    
                    ws18.cell(row=start_row+1, column=1, value="Phrase")
                    ws18.cell(row=start_row+1, column=2, value="Count")
                    ws18.cell(row=start_row+1, column=1).font = Font(bold=True, color="FFFFFF")
                    ws18.cell(row=start_row+1, column=2).font = Font(bold=True, color="FFFFFF")
                    ws18.cell(row=start_row+1, column=1).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                    ws18.cell(row=start_row+1, column=2).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                        
                    for i, (_, row) in enumerate(trigrams.iterrows(), start_row+2):
                        ws18.cell(row=i, column=1, value=row['phrase'])
                        ws18.cell(row=i, column=2, value=row['count'])
    
                    # Trigram Chart
                    tri_chart = trigrams.head(15).sort_values('count', ascending=True)
                    fig_tri = px.bar(
                        tri_chart, y='phrase', x='count',
                        orientation='h',
                        color='count',
                        color_continuous_scale=['#FF9800', '#F44336'],
                        title='Top 15 Trigrams'
                    )
                    fig_tri.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=400, showlegend=False)
                    img_bytes = fig_to_image_bytes(fig_tri, 500, 400)
                    img = XLImage(img_bytes)
                    ws18.add_image(img, f'D{start_row}')

                # ==========================================
                # SHEET 19: KEYWORDS BY RATING
                # ==========================================
                ws19 = wb.create_sheet("🎯 Keywords by Rating")

                ws19['A1'] = "Distinctive Keywords by Rating Level"
                ws19['A1'].font = Font(bold=True, size=16)

                keywords_by_rating = analyzer.get_keywords_by_rating(top_n=20)

                if keywords_by_rating:
                    # Column headers
                    col_config = [
                        ('A', '1-2 Stars (Unhappy)', 'D32F2F', 'low'),
                        ('D', '3 Stars (Neutral)', 'FF9800', 'mid'),
                        ('G', '4-5 Stars (Happy)', '4CAF50', 'high')
                    ]
    
                    for col_letter, title, color, key in col_config:
                        col_idx = ord(col_letter) - ord('A') + 1
        
                        # Title
                        ws19.cell(row=3, column=col_idx, value=title)
                        ws19.cell(row=3, column=col_idx).font = Font(bold=True, color="FFFFFF")
                        ws19.cell(row=3, column=col_idx).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        ws19.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+1)
                        
                        # Sub-headers
                        ws19.cell(row=4, column=col_idx, value="Word")
                        ws19.cell(row=4, column=col_idx+1, value="Count")
                        ws19.cell(row=4, column=col_idx).font = Font(bold=True)
                        ws19.cell(row=4, column=col_idx+1).font = Font(bold=True)
        
                        # Data
                        if key in keywords_by_rating and keywords_by_rating[key]:
                            for i, item in enumerate(keywords_by_rating[key][:20], 5):
                                ws19.cell(row=i, column=col_idx, value=item['word'])
                                ws19.cell(row=i, column=col_idx+1, value=item['count'])

                # ==========================================
                # SHEET 20: TOPIC DISCOVERY
                # ==========================================
                ws20 = wb.create_sheet("🏷️ Topics")

                ws20['A1'] = "Topic Discovery - Main Themes in Comments"
                ws20['A1'].font = Font(bold=True, size=16)

                topics = analyzer.get_topic_keywords(n_topics=5, n_words=10)

                if topics:
                    # Headers
                    ws20['A3'] = "Topic"
                    ws20['B3'] = "Mention Count"
                    ws20['C3'] = "Top Keywords"
                    for col in ['A3', 'B3', 'C3']:
                        ws20[col].font = Font(bold=True, color="FFFFFF")
                        ws20[col].fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
    
                    # Data
                    for i, topic in enumerate(topics, 4):
                        ws20.cell(row=i, column=1, value=topic['topic'])
                        ws20.cell(row=i, column=2, value=topic['count'])
                        ws20.cell(row=i, column=3, value=', '.join(topic['keywords']))
    
                    # Topic Chart
                    df_topics = pd.DataFrame(topics)
                    fig_topic = px.bar(
                        df_topics.sort_values('count', ascending=True),
                        y='topic', x='count',
                        orientation='h',
                        color='count',
                        color_continuous_scale=['#9C27B0', '#E91E63'],
                        title='Topics by Mention Count'
                    )
                    fig_topic.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=350, showlegend=False)
                    img_bytes = fig_to_image_bytes(fig_topic, 500, 350)
                    img = XLImage(img_bytes)
                    ws20.add_image(img, 'E3')

                # ==========================================
                # SHEET 21: SENTIMENT ANALYSIS
                # ==========================================
                ws21 = wb.create_sheet("😊 Sentiment")
                
                ws21['A1'] = "Sentiment Analysis"
                ws21['A1'].font = Font(bold=True, size=16)

                # Sentiment Distribution
                ws21['A3'] = "Sentiment Distribution"
                ws21['A3'].font = Font(bold=True, size=14)

                sentiment_dist = analyzer.get_comment_sentiment_distribution()

                if len(sentiment_dist) > 0:
                    ws21['A4'] = "Sentiment"
                    ws21['B4'] = "Count"
                    ws21['C4'] = "Percentage"
                    ws21['D4'] = "Avg Rating"
                    for col in ['A4', 'B4', 'C4', 'D4']:
                        ws21[col].font = Font(bold=True, color="FFFFFF")
                        ws21[col].fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
                        
                    sentiment_colors = {'positive': '4CAF50', 'negative': 'D32F2F', 'neutral': '9E9E9E', 'mixed': 'FF9800'}
    
                    for i, (_, row) in enumerate(sentiment_dist.iterrows(), 5):
                        ws21.cell(row=i, column=1, value=row['sentiment'])
                        ws21.cell(row=i, column=2, value=row['count'])
                        ws21.cell(row=i, column=3, value=f"{row['percentage']}%")
                        ws21.cell(row=i, column=4, value=round(row['avg_rating'], 2))
        
                        # Color code sentiment
                        color = sentiment_colors.get(row['sentiment'], '9E9E9E')
                        ws21.cell(row=i, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        ws21.cell(row=i, column=1).font = Font(color="FFFFFF")
    
                    # Sentiment Pie Chart
                    fig_sent = px.pie(
                        sentiment_dist,
                        values='count',
                        names='sentiment',
                        color='sentiment',
                        color_discrete_map={'positive': '#4CAF50', 'negative': '#D32F2F', 'neutral': '#9E9E9E', 'mixed': '#FF9800'},
                        title='Sentiment Distribution'
                    )
                    fig_sent.update_layout(paper_bgcolor='white', height=350)
                    img_bytes = fig_to_image_bytes(fig_sent, 450, 350)
                    img = XLImage(img_bytes)
                    ws21.add_image(img, 'F3')

                # Rating vs Sentiment Matrix
                rating_sentiment = analyzer.get_rating_sentiment_matrix()

                if len(rating_sentiment) > 0:
                    start_row = 12
    
                    ws21.cell(row=start_row, column=1, value="Rating vs Sentiment Matrix")
                    ws21.cell(row=start_row, column=1).font = Font(bold=True, size=14)
    
                    # Write matrix
                    for c_idx, col_name in enumerate(rating_sentiment.columns, 2):
                        ws21.cell(row=start_row+1, column=c_idx, value=col_name)
                        ws21.cell(row=start_row+1, column=c_idx).font = Font(bold=True)
                        
                        ws21.cell(row=start_row+1, column=1, value="Rating")
                        ws21.cell(row=start_row+1, column=1).font = Font(bold=True)
    
                    for r_idx, (rating, row_data) in enumerate(rating_sentiment.iterrows(), start_row+2):
                        ws21.cell(row=r_idx, column=1, value=rating)
                        for c_idx, value in enumerate(row_data, 2):
                            ws21.cell(row=r_idx, column=c_idx, value=value)

                # ==========================================
                # SHEET 22: TEXT MINING SUMMARY
                # ==========================================
                ws22 = wb.create_sheet("📊 Text Mining Summary")
                
                ws22['A1'] = "Text Mining Summary & Insights"
                ws22['A1'].font = Font(bold=True, size=16)

                ws22['A3'] = "Key Insights"
                ws22['A3'].font = Font(bold=True, size=14)

                insights = []

                # Top word
                if word_freq:
                    top_word = list(word_freq.keys())[0]
                    top_count = word_freq[top_word]
                    insights.append(f"Most frequent word: '{top_word}' ({top_count} mentions)")

                # Top bigram
                if len(bigrams) > 0:
                    top_phrase = bigrams.iloc[0]['phrase']
                    phrase_count = bigrams.iloc[0]['count']
                    insights.append(f"Most common phrase: '{top_phrase}' ({phrase_count} mentions)")

                # Top topic
                if topics:
                    top_topic = topics[0]['topic']
                    topic_count = topics[0]['count']
                    insights.append(f"Main topic: '{top_topic}' ({topic_count} mentions)")

                # Sentiment
                if len(sentiment_dist) > 0:
                    positive_pct = sentiment_dist[sentiment_dist['sentiment'] == 'positive']['percentage'].values
                    negative_pct = sentiment_dist[sentiment_dist['sentiment'] == 'negative']['percentage'].values
                    if len(positive_pct) > 0:
                        insights.append(f"Positive sentiment: {positive_pct[0]}%")
                    if len(negative_pct) > 0:
                        insights.append(f"Negative sentiment: {negative_pct[0]}%")

                # Keywords insight
                if keywords_by_rating:
                    if 'low' in keywords_by_rating and keywords_by_rating['low']:
                        low_words = [item['word'] for item in keywords_by_rating['low'][:3]]
                        insights.append(f"Unhappy customers mention: {', '.join(low_words)}")
                    if 'high' in keywords_by_rating and keywords_by_rating['high']:
                        high_words = [item['word'] for item in keywords_by_rating['high'][:3]]
                        insights.append(f"Happy customers mention: {', '.join(high_words)}")

                # Write insights
                for i, insight in enumerate(insights, 4):
                    ws22.cell(row=i, column=1, value=f"• {insight}")

                # Recommendations
                ws22.cell(row=len(insights)+6, column=1, value="Recommendations Based on Text Analysis")
                ws22.cell(row=len(insights)+6, column=1).font = Font(bold=True, size=14)

                recommendations = [
    "1. Address the most frequent negative phrases in customer training",
    "2. Highlight positive keywords in marketing materials",
    "3. Create targeted responses for each topic category",
    "4. Monitor sentiment trends over time",
    "5. Focus on converting neutral sentiment to positive"
                ]

                for i, rec in enumerate(recommendations, len(insights)+7):
                    ws22.cell(row=i, column=1, value=rec)
                
                # ==========================================
                # SHEET 23: ML - DETRACTOR PREDICTION
                # ==========================================
                ws23 = wb.create_sheet("🎯 ML Detractor Prediction")
                
                ws23['A1'] = "Machine Learning: Detractor Prediction Model"
                ws23['A1'].font = Font(bold=True, size=16)

                # Initialize ML Analyzer
                from ml_analyzer import ShilaMLAnalyzer
                from config import COLS
                ml_analyzer = ShilaMLAnalyzer(df, COLS)

                try:
                    detractor_results = ml_analyzer.train_detractor_model()
                    
                    if 'error' not in detractor_results:
                        # Model Performance
                        ws23['A3'] = "Model Performance"
                        ws23['A3'].font = Font(bold=True, size=14)
        
                        metrics = [
                            ('Metric', 'Value'),
                            ('Accuracy', f"{detractor_results['accuracy']*100:.1f}%"),
                            ('Precision', f"{detractor_results['precision']*100:.1f}%"),
                            ('Recall', f"{detractor_results['recall']*100:.1f}%"),
                            ('F1 Score', f"{detractor_results['f1_score']*100:.1f}%"),
                            ('Cross-Val Mean', f"{detractor_results['cv_mean']*100:.1f}%"),
                            ('Train Size', detractor_results['train_size']),
                            ('Test Size', detractor_results['test_size']),
                            ('Detractor Rate', f"{detractor_results['detractor_rate']}%")
                        ]
        
                    for i, (metric, value) in enumerate(metrics, 4):
                        ws23.cell(row=i, column=1, value=metric)
                        ws23.cell(row=i, column=2, value=value)
                        if i == 4:
                            ws23.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
                            ws23.cell(row=i, column=2).font = Font(bold=True, color="FFFFFF")
                            ws23.cell(row=i, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                            ws23.cell(row=i, column=2).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                        
                        # Feature Importance
                        ws23['A15'] = "Feature Importance"
                        ws23['A15'].font = Font(bold=True, size=14)
                        
                        ws23['A16'] = "Feature"
                        ws23['B16'] = "Importance"
                        ws23['A16'].font = Font(bold=True, color="FFFFFF")
                        ws23['B16'].font = Font(bold=True, color="FFFFFF")
                        ws23['A16'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                        ws23['B16'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                
                    for i, feat in enumerate(detractor_results['feature_importance'][:10], 17):
                        ws23.cell(row=i, column=1, value=feat['feature'])
                        ws23.cell(row=i, column=2, value=round(feat['importance'], 4))
        
                    # Confusion Matrix
                    ws23['A30'] = "Confusion Matrix"
                    ws23['A30'].font = Font(bold=True, size=14)
                    
                    cm = detractor_results['confusion_matrix']
                    ws23['B32'] = "Predicted: No"
                    ws23['C32'] = "Predicted: Yes"
                    ws23['A33'] = "Actual: No"
                    ws23['A34'] = "Actual: Yes"
                    ws23['B33'] = cm[0][0]
                    ws23['C33'] = cm[0][1]
                    ws23['B34'] = cm[1][0]
                    ws23['C34'] = cm[1][1]
                
                    # Feature Importance Chart
                    feat_df = pd.DataFrame(detractor_results['feature_importance'][:10])
                    feat_df = feat_df.sort_values('importance', ascending=True)
                    
                    fig_feat = px.bar(
                        feat_df, y='feature', x='importance',
                        orientation='h',
                        color='importance',
                        color_continuous_scale=['#FFC107', '#4CAF50'],
                        title='Feature Importance for Detractor Prediction'
                        )
                    fig_feat.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=400, showlegend=False)
                    img_bytes = fig_to_image_bytes(fig_feat, 500, 400)
                    img = XLImage(img_bytes)
                    ws23.add_image(img, 'E3')
                
                    # High Risk Customers
                    ws23['A38'] = "High Risk Customers (Top 30)"
                    ws23['A38'].font = Font(bold=True, size=14)
                
                    high_risk = ml_analyzer.predict_detractor_risk(top_n=30)
                    if len(high_risk) > 0:
                        for c_idx, col_name in enumerate(high_risk.columns, 1):
                            cell = ws23.cell(row=39, column=c_idx, value=col_name)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                        
                        for r_idx, (_, row_data) in enumerate(high_risk.iterrows(), 40):
                            for c_idx, value in enumerate(row_data, 1):
                                if isinstance(value, float):
                                    ws23.cell(row=r_idx, column=c_idx, value=round(value, 3))
                                else:
                                    ws23.cell(row=r_idx, column=c_idx, value=value)
                        else:
                            ws23['A3'] = f"Error: {detractor_results['error']}"
                except Exception as e:
                    ws23['A3'] = f"ML Analysis Error: {str(e)}"

                # ==========================================
                # SHEET 24: ML - CUSTOMER CLUSTERING
                # ==========================================
                ws24 = wb.create_sheet("👥 ML Clustering")
                
                ws24['A1'] = "Machine Learning: Customer Clustering"
                ws24['A1'].font = Font(bold=True, size=16)

                try:
                    cluster_results = ml_analyzer.perform_clustering(n_clusters=5)
    
                    if 'error' not in cluster_results:
                        # Cluster Profiles
                        ws24['A3'] = "Cluster Profiles"
                        ws24['A3'].font = Font(bold=True, size=14)
                        
                        cluster_df = pd.DataFrame(cluster_results['cluster_stats'])
                        
                        # Headers
                        headers = ['Cluster', 'Name', 'Size', 'Percentage', 'Avg Rating', 'Avg NPS', 'Promoter %', 'Detractor %']
                        for c_idx, header in enumerate(headers, 1):
                            cell = ws24.cell(row=4, column=c_idx, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        
                        # Data
                        for r_idx, (_, row) in enumerate(cluster_df.iterrows(), 5):
                            ws24.cell(row=r_idx, column=1, value=row.get('cluster', ''))
                            ws24.cell(row=r_idx, column=2, value=row.get('cluster_name', ''))
                            ws24.cell(row=r_idx, column=3, value=row.get('size', 0))
                            ws24.cell(row=r_idx, column=4, value=f"{row.get('percentage', 0)}%")
                            ws24.cell(row=r_idx, column=5, value=row.get('avg_rating', ''))
                            ws24.cell(row=r_idx, column=6, value=row.get('avg_nps', ''))
                            ws24.cell(row=r_idx, column=7, value=f"{row.get('promoter_pct', 0)}%")
                            ws24.cell(row=r_idx, column=8, value=f"{row.get('detractor_pct', 0)}%")
                            
                        # Cluster Size Chart
                        fig_cluster = px.pie(
                            cluster_df,
                            values='size',
                            names='cluster_name',
                            title='Customer Cluster Distribution',
                            color_discrete_sequence=['#4CAF50', '#8BC34A', '#FFC107', '#FF9800', '#F44336']
                        )
                        fig_cluster.update_layout(paper_bgcolor='white', height=400)
                        img_bytes = fig_to_image_bytes(fig_cluster, 500, 400)
                        img = XLImage(img_bytes)
                        ws24.add_image(img, 'J3')
        
                        # Cluster Comparison Chart
                        fig_compare = px.bar(
                            cluster_df,
                            x='cluster_name',
                            y=['avg_rating', 'avg_nps'],
                            barmode='group',
                            title='Cluster Comparison: Rating vs NPS',
                            color_discrete_sequence=['#4CAF50', '#2196F3']
                        )
                        fig_compare.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=350)
                        img_bytes = fig_to_image_bytes(fig_compare, 500, 350)
                        img = XLImage(img_bytes)
                        ws24.add_image(img, 'J22')
                    else:
                        ws24['A3'] = f"Error: {cluster_results['error']}"
                except Exception as e:
                    ws24['A3'] = f"Clustering Error: {str(e)}"

                # ==========================================
                # SHEET 25: ML - ASSOCIATION RULES
                # ==========================================
                ws25 = wb.create_sheet("🔗 ML Association Rules")
                
                ws25['A1'] = "Machine Learning: Association Rules"
                ws25['A1'].font = Font(bold=True, size=16)
                ws25['A2'] = "Which issues frequently occur together"
                ws25['A2'].font = Font(italic=True, color="666666")

                try:
                    rules_results = ml_analyzer.get_association_rules(min_support=0.01, min_confidence=0.3)
    
                    if 'error' not in rules_results:
                        # Summary
                        ws25['A4'] = f"Total Transactions: {rules_results['total_transactions']:,}"
                        ws25['A5'] = f"Unique Items: {rules_results['unique_items']}"
                        ws25['A6'] = f"Rules Found: {len(rules_results['rules'])}"
                        
                        # Association Rules Table
                        ws25['A8'] = "Association Rules"
                        ws25['A8'].font = Font(bold=True, size=14)
                        
                        rule_headers = ['IF (Antecedent)', 'THEN (Consequent)', 'Support', 'Confidence', 'Lift']
                        for c_idx, header in enumerate(rule_headers, 1):
                            cell = ws25.cell(row=9, column=c_idx, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                            
                        for r_idx, rule in enumerate(rules_results['rules'], 10):
                            ws25.cell(row=r_idx, column=1, value=rule['if'])
                            ws25.cell(row=r_idx, column=2, value=rule['then'])
                            ws25.cell(row=r_idx, column=3, value=f"{rule['support']:.1%}")
                            ws25.cell(row=r_idx, column=4, value=f"{rule['confidence']:.1%}")
                            ws25.cell(row=r_idx, column=5, value=rule['lift'])
            
                            # Color code lift
                            lift_cell = ws25.cell(row=r_idx, column=5)
                            if rule['lift'] > 1.5:
                                lift_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            elif rule['lift'] < 1:
                                lift_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
                        # Frequent Itemsets
                        start_row = 10 + len(rules_results['rules']) + 3
                        ws25.cell(row=start_row, column=1, value="Frequent Issue Combinations")
                        ws25.cell(row=start_row, column=1).font = Font(bold=True, size=14)
                        
                        item_headers = ['Items', 'Support', 'Count']
                        for c_idx, header in enumerate(item_headers, 1):
                            cell = ws25.cell(row=start_row+1, column=c_idx, value=header)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
                            
                        for r_idx, item in enumerate(rules_results['frequent_itemsets'], start_row+2):
                            ws25.cell(row=r_idx, column=1, value=item['items'])
                            ws25.cell(row=r_idx, column=2, value=f"{item['support']:.1%}")
                            ws25.cell(row=r_idx, column=3, value=item['count'])
                    else:
                        ws25['A4'] = f"Error: {rules_results['error']}"
                except Exception as e:
                    ws25['A4'] = f"Association Rules Error: {str(e)}"

                # ==========================================
                # SHEET 26: ML - ANOMALY DETECTION
                # ==========================================
                ws26 = wb.create_sheet("🚨 ML Anomaly Detection")
                
                ws26['A1'] = "Machine Learning: Anomaly Detection"
                ws26['A1'].font = Font(bold=True, size=16)
                ws26['A2'] = "Find unusual patterns in customer feedback"
                ws26['A2'].font = Font(italic=True, color="666666")

                try:
                    anomaly_results = ml_analyzer.detect_anomalies(contamination=0.05)
    
                    if 'error' not in anomaly_results:
                        stats = anomaly_results['stats']
        
                        # Summary Metrics
                        ws26['A4'] = "Summary"
                        ws26['A4'].font = Font(bold=True, size=14)
        
                        summary_data = [
                            ('Metric', 'Value'),
                            ('Total Anomalies', stats['total_anomalies']),
                            ('Anomaly Rate', f"{stats['anomaly_rate']}%"),
                            ('Anomaly Avg Rating', stats.get('anomaly_avg_rating', 'N/A')),
                            ('Normal Avg Rating', stats.get('normal_avg_rating', 'N/A')),
                            ('Anomaly Avg NPS', stats.get('anomaly_avg_nps', 'N/A')),
                            ('Normal Avg NPS', stats.get('normal_avg_nps', 'N/A'))
                        ]
        
                        for i, (metric, value) in enumerate(summary_data, 5):
                            ws26.cell(row=i, column=1, value=metric)
                            ws26.cell(row=i, column=2, value=value)
                            if i == 5:
                                ws26.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
                                ws26.cell(row=i, column=2).font = Font(bold=True, color="FFFFFF")
                                ws26.cell(row=i, column=1).fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                                ws26.cell(row=i, column=2).fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                                
                        # Anomaly Types
                        if anomaly_results['anomaly_types']:
                            ws26['A14'] = "Anomaly Types Detected"
                            ws26['A14'].font = Font(bold=True, size=14)
                            
                            type_headers = ['Type', 'Description', 'Count', 'Icon']
                            for c_idx, header in enumerate(type_headers, 1):
                                cell = ws26.cell(row=15, column=c_idx, value=header)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            
                            for r_idx, atype in enumerate(anomaly_results['anomaly_types'], 16):
                                ws26.cell(row=r_idx, column=1, value=atype['type'])
                                ws26.cell(row=r_idx, column=2, value=atype['description'])
                                ws26.cell(row=r_idx, column=3, value=atype['count'])
                                ws26.cell(row=r_idx, column=4, value=atype['icon'])
        
                        # Top Anomalies
                        ws26['A22'] = "Top Anomalies to Review"
                        ws26['A22'].font = Font(bold=True, size=14)
                        
                        top_anomalies = anomaly_results['top_anomalies']
                        if top_anomalies:
                            headers = list(top_anomalies[0].keys())
                            for c_idx, header in enumerate(headers, 1):
                                cell = ws26.cell(row=23, column=c_idx, value=header)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
            
                            for r_idx, anomaly in enumerate(top_anomalies[:20], 24):
                                for c_idx, key in enumerate(headers, 1):
                                    value = anomaly.get(key, '')
                                    if isinstance(value, float):
                                        value = round(value, 3)
                                        ws26.cell(row=r_idx, column=c_idx, value=value)
                    else:
                        ws26['A4'] = f"Error: {anomaly_results['error']}"
                except Exception as e:
                    ws26['A4'] = f"Anomaly Detection Error: {str(e)}"

                # ==========================================
                # SHEET 27: ML - CHURN PREDICTION
                # ==========================================
                ws27 = wb.create_sheet("📉 ML Churn Prediction")

                ws27['A1'] = "Machine Learning: Churn Prediction"
                ws27['A1'].font = Font(bold=True, size=16)
                ws27['A2'] = "Predict which customers are likely to stop ordering"
                ws27['A2'].font = Font(italic=True, color="666666")

                try:
                    churn_results = ml_analyzer.train_churn_model()
                    
                    if 'error' not in churn_results:
                        # Model Performance
                        ws27['A4'] = "Model Performance"
                        ws27['A4'].font = Font(bold=True, size=14)
                        
                        metrics = [
                            ('Metric', 'Value'),
                            ('Accuracy', f"{churn_results['accuracy']*100:.1f}%"),
                            ('Precision', f"{churn_results['precision']*100:.1f}%"),
                            ('Recall', f"{churn_results['recall']*100:.1f}%"),
                            ('F1 Score', f"{churn_results['f1_score']*100:.1f}%"),
                            ('Churn Rate', f"{churn_results['churn_rate']}%")
                        ]
        
                        for i, (metric, value) in enumerate(metrics, 5):
                            ws27.cell(row=i, column=1, value=metric)
                            ws27.cell(row=i, column=2, value=value)
                            if i == 5:
                                ws27.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
                                ws27.cell(row=i, column=2).font = Font(bold=True, color="FFFFFF")
                                ws27.cell(row=i, column=1).fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                                ws27.cell(row=i, column=2).fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        
                        # Note about proxy model
                        ws27['A12'] = "Note: This is a proxy model based on rating/NPS/issues. True churn requires repeat customer data."
                        ws27['A12'].font = Font(italic=True, color="666666")
                        
                        # Feature Importance
                        ws27['A14'] = "Feature Importance"
                        ws27['A14'].font = Font(bold=True, size=14)
                        
                        ws27['A15'] = "Feature"
                        ws27['B15'] = "Importance"
                        ws27['A15'].font = Font(bold=True, color="FFFFFF")
                        ws27['B15'].font = Font(bold=True, color="FFFFFF")
                        ws27['A15'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                        ws27['B15'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                        
                        for i, feat in enumerate(churn_results['feature_importance'], 16):
                            ws27.cell(row=i, column=1, value=feat['feature'])
                            ws27.cell(row=i, column=2, value=round(feat['importance'], 4))
        
                        # Confusion Matrix
                        ws27['A22'] = "Confusion Matrix"
                        ws27['A22'].font = Font(bold=True, size=14)
                        
                        cm = churn_results['confusion_matrix']
                        ws27['B24'] = "Pred: Stay"
                        ws27['C24'] = "Pred: Churn"
                        ws27['A25'] = "Actual: Stay"
                        ws27['A26'] = "Actual: Churn"
                        ws27['B25'] = cm[0][0]
                        ws27['C25'] = cm[0][1]
                        ws27['B26'] = cm[1][0]
                        ws27['C26'] = cm[1][1]
                        
                        # High Churn Risk Customers
                        ws27['A30'] = "High Churn Risk Customers (Top 30)"
                        ws27['A30'].font = Font(bold=True, size=14)
                        
                        churn_risk = ml_analyzer.predict_churn_risk(top_n=30)
                        if len(churn_risk) > 0:
                            for c_idx, col_name in enumerate(churn_risk.columns, 1):
                                cell = ws27.cell(row=31, column=c_idx, value=col_name)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            
                            for r_idx, (_, row_data) in enumerate(churn_risk.iterrows(), 32):
                                for c_idx, value in enumerate(row_data, 1):
                                    if isinstance(value, float):
                                        ws27.cell(row=r_idx, column=c_idx, value=round(value, 3))
                                    else:
                                        ws27.cell(row=r_idx, column=c_idx, value=value)
        
                        # Chart
                        feat_df = pd.DataFrame(churn_results['feature_importance'])
                        feat_df = feat_df.sort_values('importance', ascending=True)
                        
                        fig_churn = px.bar(
                            feat_df, y='feature', x='importance',
                            orientation='h',
                            color='importance',
                            color_continuous_scale=['#BBDEFB', '#2196F3'],
                            title='Feature Importance for Churn Prediction'
                            )
                        fig_churn.update_layout(paper_bgcolor='white', plot_bgcolor='white', height=300, showlegend=False)
                        img_bytes = fig_to_image_bytes(fig_churn, 450, 300)
                        img = XLImage(img_bytes)
                        ws27.add_image(img, 'E4')
                    else:
                        ws27['A4'] = f"Error: {churn_results['error']}"
                except Exception as e:
                    ws27['A4'] = f"Churn Prediction Error: {str(e)}"

                # ==========================================
                # SHEET 28: ML SUMMARY
                # ==========================================
                ws28 = wb.create_sheet("📊 ML Summary")
                
                ws28['A1'] = "Machine Learning Analysis Summary"
                ws28['A1'].font = Font(bold=True, size=16)
                
                ws28['A3'] = "Models Trained"
                ws28['A3'].font = Font(bold=True, size=14)

                ml_models = [
                    ('Model', 'Purpose', 'Key Metric', 'Status'),
                    ('Detractor Prediction', 'Predict unhappy customers before they complain', 'F1 Score', '✅ Trained'),
                    ('Customer Clustering', 'Find natural customer segments', '5 Clusters', '✅ Trained'),
                    ('Association Rules', 'Find issue combinations', 'Lift Score', '✅ Trained'),
                    ('Anomaly Detection', 'Find unusual patterns', 'Anomaly Rate', '✅ Trained'),
                    ('Churn Prediction', 'Predict customer churn', 'F1 Score', '✅ Trained')
                    ]
                
                for i, row_data in enumerate(ml_models, 4):
                    for j, value in enumerate(row_data, 1):
                        cell = ws28.cell(row=i, column=j, value=value)
                        if i == 4:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="673AB7", end_color="673AB7", fill_type="solid")
                            
                ws28['A12'] = "Key Insights"
                ws28['A12'].font = Font(bold=True, size=14)

                insights = [
                    "1. Use Detractor Prediction to identify at-risk customers before they leave bad reviews",
                    "2. Customer Clustering reveals 5 distinct segments - target each with specific strategies",
                    "3. Association Rules show which issues tend to occur together - fix root causes",
                    "4. Anomaly Detection flags suspicious patterns that may indicate fraud or system errors",
                    "5. Churn Prediction helps prioritize retention efforts on high-risk customers"
                    ]

                for i, insight in enumerate(insights, 13):
                    ws28.cell(row=i, column=1, value=insight)
                    
                ws28['A20'] = "Recommendations"
                ws28['A20'].font = Font(bold=True, size=14)

                recommendations = [
                    "• Train models weekly with new data for best accuracy",
                    "• Focus retention efforts on High Risk detractor/churn customers",
                    "• Investigate anomalies promptly - they may indicate fraud",
                    "• Use clustering insights for targeted marketing campaigns",
                    "• Address issue combinations identified by association rules"
                    ]

                for i, rec in enumerate(recommendations, 21):
                    ws28.cell(row=i, column=1, value=rec)
                
                # ==========================================
                # SAVE WORKBOOK
                # ==========================================
                wb.save(fp)
                
                # Download button
                with open(fp, 'rb') as f:
                    st.download_button(
                        "📥 Click to Download Excel with Charts",
                        f,
                        f"shila_full_report_{ts}.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                st.success(f"✅ Excel report generated with 22 sheets and embedded charts!")
                
            except Exception as e:
                st.error(f"Export failed: {e}")
                st.exception(e)

with c_exp_2:
    st.button(f"📊 {L('export_pptx')}", disabled=True, use_container_width=True, help="Coming Soon")

with c_exp_3:
    if st.button(f"📝 {L('export_notebooklm')}", use_container_width=True):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Build comprehensive markdown report
        md_content = f"""# 📊 Shila Restaurant - Analysis Report
**Generated:** {ts}

---

## 📈 Key Performance Indicators
| Metric | Value |
|--------|-------|
| NPS Score | {kpis['nps_score']} |
| Average Rating | {kpis['avg_rating']} / 5 |
| Total Orders | {kpis['total_orders']:,} |
| Promoters | {kpis['promoters']:,} |
| Detractors | {kpis['detractors']:,} |
| Response Rate | {kpis['response_rate']}% |

---

## 🚨 Top Issues (Pain Points)
"""
        for _, r in analyzer.get_top_issues(10).iterrows():
            md_content += f"- **{r['Issue']}**: {r['Count']} mentions\n"
        
        md_content += "\n---\n\n## 🏆 Top Strengths\n"
        for _, r in analyzer.get_top_strengths(10).iterrows():
            md_content += f"- **{r['Strength']}**: {r['Count']} mentions\n"
        
        # Pareto Analysis
        pareto = analyzer.get_pareto_analysis()
        if len(pareto) > 0:
            md_content += "\n---\n\n## 📊 Pareto Analysis (Top 5 Issues by Impact)\n"
            md_content += "| Issue | Damage Score | Frequency | Avg Rating |\n"
            md_content += "|-------|-------------|-----------|------------|\n"
            for _, r in pareto.head(5).iterrows():
                md_content += f"| {r['tag']} | {r['total_damage']} | {r['frequency']} | {r['avg_rating']} |\n"
        
        # Branch Analysis
        br_stats, _ = analyzer.get_branch_analysis()
        if len(br_stats) > 0:
            md_content += "\n---\n\n## 🏪 Branch Performance\n"
            md_content += "### ✅ Best Performing Branches\n"
            for _, r in br_stats.head(3).iterrows():
                nps_txt = f", NPS {r['nps_score']:.1f}" if 'nps_score' in r else ""
                md_content += f"- **{r['branch']}**: Rating {r['avg_rating']:.2f}{nps_txt} ({int(r['order_count'])} orders)\n"
            md_content += "\n### ⚠️ Needs Improvement\n"
            for _, r in br_stats.tail(3).iterrows():
                nps_txt = f", NPS {r['nps_score']:.1f}" if 'nps_score' in r else ""
                md_content += f"- **{r['branch']}**: Rating {r['avg_rating']:.2f}{nps_txt} ({int(r['order_count'])} orders)\n"
        
        # Customer Segments
        recovery = analyzer.get_recovery_opportunities()
        if len(recovery) > 0:
            md_content += "\n---\n\n## 🎯 Customer Segments\n"
            md_content += "| Segment | Persian | Count | Percentage |\n"
            md_content += "|---------|---------|-------|------------|\n"
            for _, r in recovery.iterrows():
                md_content += f"| {r['emoji']} {r['segment']} | {r['segment_fa']} | {r['count']:,} | {r['percentage']}% |\n"
        
        # Issue Categories
        issue_cats = analyzer.get_issue_category_analysis()
        if len(issue_cats) > 0:
            md_content += "\n---\n\n## ⚠️ Issue Category Impact\n"
            md_content += "| Category | Issues | % of Orders | Rating Impact | Top Problems |\n"
            md_content += "|----------|--------|-------------|---------------|---------------|\n"
            for _, r in issue_cats.iterrows():
                md_content += f"| {r['category_fa']} | {r['issue_count']} | {r['issue_pct']}% | -{r['rating_impact']:.2f} | {r['top_issues']} |\n"
        
        # Month-over-Month
        mom = analyzer.get_mom_comparison()
        if len(mom) > 1:
            md_content += "\n---\n\n## 📅 Month-over-Month Trend\n"
            md_content += "| Month | Orders | Avg Rating | NPS Score |\n"
            md_content += "|-------|--------|------------|------------|\n"
            for _, r in mom.tail(6).iterrows():
                nps_txt = f"{r['nps_score']:.1f}" if 'nps_score' in r and pd.notna(r['nps_score']) else "N/A"
                md_content += f"| {r['year_month']} | {int(r['order_count'])} | {r['avg_rating']:.2f} | {nps_txt} |\n"
            
            latest = mom.iloc[-1]
            md_content += f"\n**Latest Trend:**\n"
            md_content += f"- Rating Change: {latest['rating_change']:+.2f}\n"
            md_content += f"- Orders Change: {latest['orders_change_pct']:+.1f}%\n"
            if 'nps_change' in latest and pd.notna(latest['nps_change']):
                md_content += f"- NPS Change: {latest['nps_change']:+.1f}\n"
        
        # Product Analysis
        products = analyzer.get_product_analysis()
        if len(products) > 0:
            md_content += "\n---\n\n## 🍔 Product Performance\n"
            md_content += "### ⭐ Top Rated Products\n"
            md_content += "| Rank | Product | Rating | Orders |\n"
            md_content += "|------|---------|--------|--------|\n"
            for _, r in products.head(5).iterrows():
                md_content += f"| {int(r['rank'])} | {r['product']} | {r['avg_rating']:.2f} | {int(r['order_count'])} |\n"
            
            if len(products) > 5:
                md_content += "\n### 📉 Needs Attention\n"
                md_content += "| Product | Rating | Orders |\n"
                md_content += "|---------|--------|--------|\n"
                for _, r in products.tail(3).iterrows():
                    md_content += f"| {r['product']} | {r['avg_rating']:.2f} | {int(r['order_count'])} |\n"
        
        # Aspect Sentiment
        aspects = analyzer.get_aspect_sentiment()
        if len(aspects) > 0:
            md_content += "\n---\n\n## 🎭 Aspect Sentiment Analysis\n"
            md_content += "| Aspect | Mentions | Avg Rating | Positive % | Negative % | Sentiment |\n"
            md_content += "|--------|----------|------------|------------|------------|------------|\n"
            for _, r in aspects.iterrows():
                sentiment = "🟢" if r['sentiment_score'] > 0.3 else ("🔴" if r['sentiment_score'] < -0.3 else "🟡")
                md_content += f"| {r['aspect']} | {r['mentions']} | {r['avg_rating']} | {r['positive_pct']}% | {r['negative_pct']}% | {sentiment} {r['sentiment_score']:.2f} |\n"
        
        # Kano Analysis
        kano = analyzer.get_kano_analysis()
        if len(kano) > 0:
            md_content += "\n---\n\n## 🎨 Kano Model Classification\n"
            for ktype in ['Must-Be', 'Performance', 'Delighter']:
                ktype_data = kano[kano['kano_type'] == ktype]
                if len(ktype_data) > 0:
                    emoji = "🔴" if ktype == "Must-Be" else ("🟡" if ktype == "Performance" else "🟢")
                    md_content += f"\n### {emoji} {ktype} Attributes\n"
                    for _, r in ktype_data.iterrows():
                        md_content += f"- **{r['attribute']}**: Lift {r['lift_as_strength']:+.2f}, Drop {r['drop_as_weakness']:+.2f}\n"
        
        # ==========================================
        # TEXT MINING SECTION
        # ==========================================
        md_content += "\n---\n\n## 📝 Text Mining Analysis\n"
        
        # Word Frequency
        word_freq = analyzer.get_word_frequency(min_freq=5, top_n=20)
        if word_freq:
            md_content += "\n### ☁️ Most Frequent Words\n"
            md_content += "| Word | Count |\n|------|-------|\n"
            for word, count in list(word_freq.items())[:15]:
                md_content += f"| {word} | {count} |\n"
        
        # Bigrams
        bigrams = analyzer.get_ngram_analysis(n=2, min_freq=3, top_n=15)
        if len(bigrams) > 0:
            md_content += "\n### 🔗 Common Phrases (Bigrams)\n"
            md_content += "| Phrase | Count |\n|--------|-------|\n"
            for _, row in bigrams.head(10).iterrows():
                md_content += f"| {row['phrase']} | {row['count']} |\n"
        
        # Keywords by Rating
        keywords_by_rating = analyzer.get_keywords_by_rating(top_n=10)
        if keywords_by_rating:
            md_content += "\n### 🎯 Keywords by Rating Level\n"
            
            if 'low' in keywords_by_rating and keywords_by_rating['low']:
                low_words = [item['word'] for item in keywords_by_rating['low'][:8]]
                md_content += f"\n**⭐ 1-2 Stars (Unhappy customers say):** {', '.join(low_words)}\n"
            
            if 'mid' in keywords_by_rating and keywords_by_rating['mid']:
                mid_words = [item['word'] for item in keywords_by_rating['mid'][:8]]
                md_content += f"\n**⭐⭐⭐ 3 Stars (Neutral customers say):** {', '.join(mid_words)}\n"
            
            if 'high' in keywords_by_rating and keywords_by_rating['high']:
                high_words = [item['word'] for item in keywords_by_rating['high'][:8]]
                md_content += f"\n**⭐⭐⭐⭐⭐ 4-5 Stars (Happy customers say):** {', '.join(high_words)}\n"
        
        # Topic Discovery
        topics = analyzer.get_topic_keywords(n_topics=5, n_words=8)
        if topics:
            md_content += "\n### 🏷️ Main Topics Discovered\n"
            md_content += "| Topic | Mentions | Top Keywords |\n|-------|----------|---------------|\n"
            for topic in topics:
                keywords_str = ', '.join(topic['keywords'][:5])
                md_content += f"| {topic['topic']} | {topic['count']} | {keywords_str} |\n"
        
        # Sentiment Analysis
        sentiment_dist = analyzer.get_comment_sentiment_distribution()
        if len(sentiment_dist) > 0:
            md_content += "\n### 😊 Sentiment Distribution\n"
            md_content += "| Sentiment | Count | Percentage | Avg Rating |\n"
            md_content += "|-----------|-------|------------|------------|\n"
            for _, row in sentiment_dist.iterrows():
                emoji = "🟢" if row['sentiment'] == 'positive' else ("🔴" if row['sentiment'] == 'negative' else "🟡")
                md_content += f"| {emoji} {row['sentiment']} | {row['count']:,} | {row['percentage']}% | {row['avg_rating']:.2f} |\n"
        
        # Text Mining Insights Summary
        md_content += "\n### 💡 Text Mining Key Insights\n"
        
        if word_freq:
            top_word = list(word_freq.keys())[0]
            md_content += f"- **Most mentioned word:** '{top_word}' ({word_freq[top_word]} times)\n"
        
        if len(bigrams) > 0:
            top_phrase = bigrams.iloc[0]['phrase']
            md_content += f"- **Most common phrase:** '{top_phrase}' ({bigrams.iloc[0]['count']} times)\n"
        
        if topics:
            md_content += f"- **Main discussion topic:** '{topics[0]['topic']}' ({topics[0]['count']} mentions)\n"
        
        if len(sentiment_dist) > 0:
            pos_row = sentiment_dist[sentiment_dist['sentiment'] == 'positive']
            neg_row = sentiment_dist[sentiment_dist['sentiment'] == 'negative']
            if len(pos_row) > 0:
                md_content += f"- **Positive sentiment:** {pos_row.iloc[0]['percentage']}% of comments\n"
            if len(neg_row) > 0:
                md_content += f"- **Negative sentiment:** {neg_row.iloc[0]['percentage']}% of comments\n"
        
        # ==========================================
        # MACHINE LEARNING SECTION
        # ==========================================
        md_content += "\n---\n\n## 🤖 Machine Learning Analysis\n"
        
        # Initialize ML Analyzer
        from ml_analyzer import ShilaMLAnalyzer
        from config import COLS
        ml_analyzer = ShilaMLAnalyzer(df, COLS)
        
        # Detractor Prediction
        md_content += "\n### 🎯 Detractor Prediction Model\n"
        
        try:
            detractor_results = ml_analyzer.train_detractor_model()
            
            if 'error' not in detractor_results:
                md_content += "**Model Performance:**\n"
                md_content += "| Metric | Value |\n|--------|-------|\n"
                md_content += f"| Accuracy | {detractor_results['accuracy']*100:.1f}% |\n"
                md_content += f"| Precision | {detractor_results['precision']*100:.1f}% |\n"
                md_content += f"| Recall | {detractor_results['recall']*100:.1f}% |\n"
                md_content += f"| F1 Score | {detractor_results['f1_score']*100:.1f}% |\n"
                md_content += f"| Detractor Rate | {detractor_results['detractor_rate']}% |\n"
                
                md_content += "\n**Top Features (What Predicts Detractors):**\n"
                for feat in detractor_results['feature_importance'][:5]:
                    md_content += f"- {feat['feature']}: {feat['importance']:.3f}\n"
                
                # High risk customers summary
                high_risk = ml_analyzer.predict_detractor_risk(top_n=10)
                if len(high_risk) > 0:
                    high_count = len(high_risk[high_risk['risk_level'] == 'High'])
                    md_content += f"\n**⚠️ High Risk Customers:** {high_count} customers identified as high detractor risk\n"
            else:
                md_content += f"*Model training error: {detractor_results['error']}*\n"
        except Exception as e:
            md_content += f"*Could not train detractor model: {str(e)}*\n"
        
        # Customer Clustering
        md_content += "\n### 👥 Customer Clustering\n"
        
        try:
            cluster_results = ml_analyzer.perform_clustering(n_clusters=5)
            
            if 'error' not in cluster_results:
                md_content += "**Cluster Profiles:**\n"
                md_content += "| Cluster | Size | % | Avg Rating | Avg NPS |\n"
                md_content += "|---------|------|---|------------|--------|\n"
                
                for cluster in cluster_results['cluster_stats']:
                    md_content += f"| {cluster.get('cluster_name', 'N/A')} | {cluster['size']:,} | {cluster['percentage']}% | {cluster.get('avg_rating', 'N/A')} | {cluster.get('avg_nps', 'N/A')} |\n"
                
                md_content += "\n**Cluster Insights:**\n"
                md_content += "- Champions: Highest rating & NPS - your best customers\n"
                md_content += "- At Risk/Critical: Need immediate attention and recovery efforts\n"
            else:
                md_content += f"*Clustering error: {cluster_results['error']}*\n"
        except Exception as e:
            md_content += f"*Could not perform clustering: {str(e)}*\n"
        
        # Association Rules
        md_content += "\n### 🔗 Association Rules (Issue Combinations)\n"
        
        try:
            rules_results = ml_analyzer.get_association_rules(min_support=0.01, min_confidence=0.3)
            
            if 'error' not in rules_results:
                md_content += f"*Found {len(rules_results['rules'])} rules from {rules_results['total_transactions']:,} transactions*\n\n"
                
                md_content += "**Top Association Rules:**\n"
                md_content += "| IF (Issue) | THEN (Also Occurs) | Confidence | Lift |\n"
                md_content += "|------------|-------------------|------------|------|\n"
                
                for rule in rules_results['rules'][:8]:
                    lift_emoji = "🟢" if rule['lift'] > 1.5 else ("🟡" if rule['lift'] > 1 else "🔴")
                    md_content += f"| {rule['if']} | {rule['then']} | {rule['confidence']:.0%} | {lift_emoji} {rule['lift']:.2f} |\n"
                
                md_content += "\n**Interpretation:** Lift > 1.5 means issues strongly co-occur. Fix one to potentially fix both.\n"
            else:
                md_content += f"*Association rules error: {rules_results['error']}*\n"
        except Exception as e:
            md_content += f"*Could not mine association rules: {str(e)}*\n"
        
        # Anomaly Detection
        md_content += "\n### 🚨 Anomaly Detection\n"
        
        try:
            anomaly_results = ml_analyzer.detect_anomalies(contamination=0.05)
            
            if 'error' not in anomaly_results:
                stats = anomaly_results['stats']
                
                md_content += "**Anomaly Summary:**\n"
                md_content += f"- Total Anomalies Found: **{stats['total_anomalies']}** ({stats['anomaly_rate']}% of data)\n"
                
                if 'anomaly_avg_rating' in stats:
                    md_content += f"- Anomaly Avg Rating: {stats['anomaly_avg_rating']} vs Normal: {stats['normal_avg_rating']}\n"
                
                if anomaly_results['anomaly_types']:
                    md_content += "\n**Anomaly Types:**\n"
                    for atype in anomaly_results['anomaly_types']:
                        md_content += f"- {atype['icon']} **{atype['type']}**: {atype['count']} cases - {atype['description']}\n"
                
                md_content += "\n**Action:** Review flagged anomalies for potential fraud, system errors, or data quality issues.\n"
            else:
                md_content += f"*Anomaly detection error: {anomaly_results['error']}*\n"
        except Exception as e:
            md_content += f"*Could not detect anomalies: {str(e)}*\n"
        
        # Churn Prediction
        md_content += "\n### 📉 Churn Prediction\n"
        
        try:
            churn_results = ml_analyzer.train_churn_model()
            
            if 'error' not in churn_results:
                md_content += "*Note: Proxy model based on rating/NPS/issues. True churn requires repeat customer data.*\n\n"
                
                md_content += "**Model Performance:**\n"
                md_content += "| Metric | Value |\n|--------|-------|\n"
                md_content += f"| Accuracy | {churn_results['accuracy']*100:.1f}% |\n"
                md_content += f"| Precision | {churn_results['precision']*100:.1f}% |\n"
                md_content += f"| Recall | {churn_results['recall']*100:.1f}% |\n"
                md_content += f"| Estimated Churn Rate | {churn_results['churn_rate']}% |\n"
                
                md_content += "\n**Key Churn Drivers:**\n"
                for feat in churn_results['feature_importance']:
                    md_content += f"- {feat['feature']}: {feat['importance']:.3f}\n"
                
                # High churn risk summary
                churn_risk = ml_analyzer.predict_churn_risk(top_n=10)
                if len(churn_risk) > 0:
                    high_churn = len(churn_risk[churn_risk['churn_level'] == 'High'])
                    md_content += f"\n**⚠️ High Churn Risk:** {high_churn} customers at high risk of churning\n"
            else:
                md_content += f"*Churn model error: {churn_results['error']}*\n"
        except Exception as e:
            md_content += f"*Could not train churn model: {str(e)}*\n"
        
        # ML Summary & Recommendations
        md_content += "\n### 💡 ML-Driven Recommendations\n"
        md_content += """
1. **Proactive Outreach**: Contact high detractor-risk customers before they complain
2. **Segment Marketing**: Tailor campaigns to each customer cluster
3. **Root Cause Analysis**: Fix issue combinations identified by association rules
4. **Fraud Prevention**: Investigate anomalies flagged by detection model
5. **Retention Focus**: Prioritize high churn-risk customers for loyalty programs
6. **Continuous Learning**: Retrain models monthly with new data

**Data Quality Note:** ML models perform best with:
- Customer ID for tracking repeat behavior
- Order timestamps for temporal patterns
- Complete feedback data (minimize missing values)
"""
              
        # Footer
        md_content += "\n---\n\n## 📝 Notes for Analysis\n"
        md_content += "- **NPS Score** ranges from -100 to +100 (yours: " + str(kpis['nps_score']) + ")\n"
        md_content += "- **Pareto Principle**: Focus on top issues that cause 80% of damage\n"
        md_content += "- **Kano Types**: Must-Be (fix first), Performance (improve), Delighter (innovate)\n"
        md_content += "- **Customer Segments**: Prioritize 'At Risk' and 'Silent Churners'\n"
        md_content += "\n---\n\n*Report generated by InsightForge QFD Dashboard*\n"
        
        # Save file
        fp = os.path.join(NOTEBOOKLM_DIR, f"shila_report_{ts}.md")
        with open(fp, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        # Download button
        with open(fp, 'rb') as f:
            st.download_button(
                "📥 Download Full Report", 
                f, 
                f"shila_report_{ts}.md", 
                "text/markdown"
            )
        
        st.success(f"✅ Report generated with {len(md_content):,} characters!")



